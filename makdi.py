#!/usr/bin/env python3
"""
Makdi - the free unlimited SEO crawler
https://vikasdisale.com/makdi
=======================================
Tabs: Internal / External / Security / Response Codes / URL /
      Page Titles / Meta Description / Meta Keywords / H1 / H2 /
      Content / Images / Canonicals / PageSpeed / Custom Search / Changes

Usage:     py seo_crawler_web.py     ->  opens http://localhost:8090
Requires:  pip install requests beautifulsoup4 lxml
"""

import hashlib
import json
import os
import re
import threading
import time
import webbrowser
import zlib
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from urllib.parse import urljoin, urlparse, urldefrag

import requests
from bs4 import BeautifulSoup

VERSION = "1.0.0"
TOOL_NAME = "Makdi"
TOOL_URL = "https://vikasdisale.com/makdi"
PORT = 8090
TIMEOUT = 15
WORKERS = 10
USER_AGENT = f"MakdiBot/{VERSION} (+{TOOL_URL})"
TITLE_MIN, TITLE_MAX = 30, 60
DESC_MIN, DESC_MAX = 70, 160
THIN_WORDS = 300
DEEP_CLICKS = 4
SLOW_MS = 2000
LARGE_KB = 1024
MANY_RESOURCES = 60
LARGE_IMG_KB = 200
URL_MAX_LEN = 115
MAX_EXTERNAL_CHECKS = 2000
MAX_RESOURCE_CHECKS = 5000
MAX_HTML_STORE_BYTES = 120 * 1024 * 1024   # cap compressed html store (custom search)

import sys
import logging

def _data_dir():
    """Script mode: next to this file. Packaged exe: Documents\\Makdi."""
    if getattr(sys, "frozen", False):
        d = os.path.join(os.path.expanduser("~"), "Documents", TOOL_NAME)
    else:
        d = os.path.dirname(os.path.abspath(__file__))
    os.makedirs(d, exist_ok=True)
    return d

DATA_DIR = _data_dir()
HISTORY_DIR = os.path.join(DATA_DIR, "crawl_history")

# File logging so packaged-exe problems are diagnosable (exe has no console)
logging.basicConfig(
    filename=os.path.join(DATA_DIR, "makdi.log"), level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s")
logging.info("Makdi v%s starting (frozen=%s)", VERSION, getattr(sys, "frozen", False))

SKIP_EXTENSIONS = (
    ".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg", ".ico", ".pdf",
    ".zip", ".rar", ".mp3", ".mp4", ".avi", ".mov", ".doc", ".docx",
    ".xls", ".xlsx", ".ppt", ".pptx", ".css", ".js", ".json", ".xml",
    ".woff", ".woff2", ".ttf", ".eot", ".webm", ".apk",
)


def normalize_url(url):
    url, _ = urldefrag(url.strip())
    p = urlparse(url)
    if p.scheme and p.netloc and p.path == "":
        url = f"{p.scheme}://{p.netloc}/" + (f"?{p.query}" if p.query else "")
    return url


def same_domain(url, root_netloc):
    netloc = urlparse(url).netloc.lower()
    root = root_netloc.lower()
    return netloc == root or netloc == "www." + root or "www." + netloc == root


def is_crawlable(url):
    path = urlparse(url).path.lower()
    return not any(path.endswith(ext) for ext in SKIP_EXTENSIONS)


class CrawlState:
    def __init__(self):
        self.lock = threading.Lock()
        self.reset()

    def reset(self):
        self.running = False
        self.stop_flag = False
        self.phase = "idle"
        self.mode = "spider"
        self.robots_found = False
        self.changes = None
        self.start_url = ""
        self.error_note = ""
        self.pages = {}
        self.redirects = []
        self.link_sources = defaultdict(set)
        self.links = []
        self.resources = {}
        self.images = {}
        self.ext_status = {}
        self.html_store = {}         # url -> zlib-compressed html (custom search)
        self.html_bytes = 0
        self.queued_count = 0
        self.assets_total = 0
        self.assets_done = 0

STATE = CrawlState()
session = requests.Session()
session.headers.update({"User-Agent": USER_AGENT})


def fetch(url, depth):
    t0 = time.time()
    try:
        resp = session.get(url, timeout=TIMEOUT, allow_redirects=True)
    except requests.RequestException as e:
        return {"url": url, "status": "ERROR", "error": str(e)[:200],
                "depth": depth, "final_url": url, "links": [], "resources": []}
    elapsed_ms = int((time.time() - t0) * 1000)

    if resp.history:
        chain = [{"url": r.url, "status": r.status_code} for r in resp.history]
        chain.append({"url": resp.url, "status": resp.status_code})
        with STATE.lock:
            STATE.redirects.append({
                "from": url, "to": resp.url, "chain": chain,
                "hops": len(resp.history),
                "first_status": resp.history[0].status_code})

    data = {"url": url, "final_url": normalize_url(resp.url),
            "status": resp.status_code, "depth": depth, "error": "",
            "ms": elapsed_ms, "kb": round(len(resp.content) / 1024, 1),
            "links": [], "resources": []}

    ctype = resp.headers.get("Content-Type", "")
    if resp.status_code == 200 and "text/html" in ctype:
        # ---- security headers ----
        h = {k.lower(): v for k, v in resp.headers.items()}
        data["sec_missing"] = [name for key, name in [
            ("strict-transport-security", "HSTS"),
            ("x-frame-options", "X-Frame-Options"),
            ("x-content-type-options", "X-Content-Type-Options"),
            ("content-security-policy", "CSP"),
        ] if key not in h]

        soup = BeautifulSoup(resp.text, "lxml")
        base = data["final_url"]
        page_https = base.startswith("https://")

        t = soup.find("title")
        data["title"] = t.get_text(strip=True) if t else ""
        d = soup.find("meta", attrs={"name": re.compile(r"^description$", re.I)})
        data["description"] = (d.get("content") or "").strip() if d else ""
        kw = soup.find("meta", attrs={"name": re.compile(r"^keywords$", re.I)})
        data["keywords"] = (kw.get("content") or "").strip() if kw else ""
        r = soup.find("meta", attrs={"name": re.compile(r"^robots$", re.I)})
        data["meta_robots"] = (r.get("content") or "").strip() if r else ""
        data["noindex"] = "noindex" in data["meta_robots"].lower()

        canon_tags = soup.find_all("link", rel=lambda v: v and "canonical" in v)
        data["canonical_count"] = len(canon_tags)
        data["canonical"] = normalize_url(canon_tags[0].get("href", "")) if canon_tags else ""

        h1s = [h1.get_text(strip=True) for h1 in soup.find_all("h1")]
        data["h1_count"] = len(h1s)
        data["h1"] = h1s[0] if h1s else ""
        data["h1_all"] = " | ".join(h1s)
        h2s = [h2.get_text(strip=True) for h2 in soup.find_all("h2")]
        data["h2_count"] = len(h2s)
        data["h2"] = h2s[0] if h2s else ""
        data["h2_all"] = " | ".join(h2s[:5])

        data["imgs_no_alt"] = sum(
            1 for img in soup.find_all("img")
            if img.get("alt") is None or img.get("alt").strip() == "")

        # ---- lang / hreflang ----
        html_tag = soup.find("html")
        data["lang"] = (html_tag.get("lang") or "").strip() if html_tag else ""
        data["hreflang_count"] = len(soup.find_all(
            "link", rel=lambda v: v and "alternate" in v, hreflang=True))

        # ---- structured data (JSON-LD) ----
        data["schema_types"] = []
        data["schema_invalid"] = 0
        for sc in soup.find_all("script", attrs={"type": re.compile(r"ld\+json", re.I)}):
            raw = sc.string or sc.get_text() or ""
            try:
                obj = json.loads(raw)
            except Exception:
                data["schema_invalid"] += 1
                continue

            def collect(o):
                if isinstance(o, dict):
                    tt = o.get("@type")
                    if isinstance(tt, str):
                        data["schema_types"].append(tt)
                    elif isinstance(tt, list):
                        data["schema_types"].extend(x for x in tt if isinstance(x, str))
                    g = o.get("@graph")
                    if isinstance(g, list):
                        for v in g:
                            collect(v)
                elif isinstance(o, list):
                    for v in o:
                        collect(v)
            collect(obj)

        # ---- links (with anchor text) ----
        mixed = 0
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
                continue
            absolute = normalize_url(urljoin(base, href))
            if absolute.startswith("http"):
                anchor = a.get_text(strip=True)[:80] or "(no anchor text)"
                data["links"].append((absolute, anchor))

        # ---- resources + image data + mixed content ----
        data["images"] = []
        for img in soup.find_all("img"):
            src = img.get("src") or img.get("data-src")
            if src and not src.startswith("data:"):
                img_url = normalize_url(urljoin(base, src.strip()))
                data["resources"].append((img_url, "image"))
                alt = img.get("alt")
                data["images"].append((
                    img_url, (alt or "").strip(),
                    (img.get("loading") or "").lower() == "lazy"))
        for lnk in soup.find_all("link", rel=lambda v: v and "stylesheet" in v):
            href = lnk.get("href")
            if href:
                data["resources"].append((normalize_url(urljoin(base, href.strip())), "css"))
        for sc in soup.find_all("script", src=True):
            data["resources"].append((normalize_url(urljoin(base, sc["src"].strip())), "js"))
        if page_https:
            mixed = sum(1 for u, _t in data["resources"] if u.startswith("http://"))
        data["mixed"] = mixed
        data["https"] = page_https

        # ---- custom-search store (compressed raw html) ----
        data["html_z"] = zlib.compress(resp.text[:400000].encode("utf-8", "ignore"), 6)

        # ---- word count + body hash (after extraction of noise) ----
        for tag in soup(["script", "style", "noscript", "template"]):
            tag.extract()
        body = soup.find("body") or soup
        text = body.get_text(" ")
        data["word_count"] = len(re.findall(r"\w+", text, re.UNICODE))
        norm = re.sub(r"\s+", " ", text).strip().lower()
        data["body_hash"] = hashlib.md5(norm.encode("utf-8", "ignore")).hexdigest() if norm else ""
    return data


def check_status(url):
    try:
        r = session.head(url, timeout=10, allow_redirects=True)
        if r.status_code in (403, 405, 501):
            r = session.get(url, timeout=10, stream=True)
            r.close()
        return r.status_code
    except requests.RequestException:
        return "ERROR"


def check_image(url):
    out = {"status": "ERROR", "kb": None, "ctype": ""}
    try:
        r = session.head(url, timeout=10, allow_redirects=True)
        if r.status_code in (403, 405, 501) or "content-length" not in r.headers:
            r = session.get(url, timeout=10, stream=True)
            r.close()
        out["status"] = r.status_code
        out["ctype"] = r.headers.get("Content-Type", "")
        cl = r.headers.get("Content-Length")
        if cl and cl.isdigit():
            out["kb"] = round(int(cl) / 1024, 1)
    except requests.RequestException:
        pass
    return out


def load_robots(scheme, netloc):
    from urllib.robotparser import RobotFileParser
    try:
        r = session.get(f"{scheme}://{netloc}/robots.txt", timeout=10)
        if r.status_code == 200 and r.text.strip():
            rp = RobotFileParser()
            rp.parse(r.text.splitlines())
            return rp, True
    except requests.RequestException:
        pass
    return None, False


def fetch_sitemap_urls(sitemap_url):
    from xml.etree import ElementTree
    urls, seen = [], set()

    def parse(sm_url):
        if sm_url in seen or len(seen) > 50:
            return
        seen.add(sm_url)
        try:
            r = session.get(sm_url, timeout=TIMEOUT)
            if r.status_code != 200:
                return
            root = ElementTree.fromstring(r.content)
        except Exception:
            return
        ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
        for loc in root.findall(".//sm:sitemap/sm:loc", ns):
            if loc.text:
                parse(loc.text.strip())
        for loc in root.findall(".//sm:url/sm:loc", ns):
            if loc.text:
                urls.append(normalize_url(loc.text.strip()))

    parse(sitemap_url)
    return urls


def crawl_worker(start_url, max_pages, workers, mode="spider"):
    root_netloc = urlparse(start_url).netloc
    scheme = urlparse(start_url).scheme or "https"
    rp, robots_found = load_robots(scheme, root_netloc)
    with STATE.lock:
        STATE.mode = mode
        STATE.robots_found = robots_found

    def robots_allowed(u):
        if rp is None:
            return True
        try:
            return rp.can_fetch(USER_AGENT, u)
        except Exception:
            return True

    if mode == "sitemap":
        sitemap_url = start_url
        if not sitemap_url.lower().endswith(".xml"):
            sitemap_url = f"{scheme}://{root_netloc}/sitemap.xml"
        sm_urls = fetch_sitemap_urls(sitemap_url)[:max_pages]
        frontier = [(u, 0) for u in sm_urls]
        queued = set(sm_urls)
        follow_links = False
        if not frontier:
            with STATE.lock:
                STATE.running = False
                STATE.phase = "done"
                STATE.error_note = f"No URLs found in sitemap: {sitemap_url}"
            return
    else:
        frontier = [(start_url, 0)]
        queued = {start_url}
        follow_links = True

    visited = set()
    with STATE.lock:
        STATE.phase = "pages"

    # ---------- Phase 1: pages ----------
    with ThreadPoolExecutor(max_workers=workers) as pool:
        while frontier and len(visited) < max_pages:
            with STATE.lock:
                if STATE.stop_flag:
                    break
            batch, frontier = frontier[:workers * 4], frontier[workers * 4:]
            futures = {}
            for url, depth in batch:
                if url in visited:
                    continue
                visited.add(url)
                allowed = robots_allowed(url)
                if not allowed and mode == "spider":
                    with STATE.lock:
                        STATE.pages[url] = {
                            "url": url, "final_url": url, "status": "BLOCKED",
                            "depth": depth, "error": "", "links": [],
                            "resources": [], "robots_blocked": True}
                    continue
                futures[pool.submit(fetch, url, depth)] = (url, depth, allowed)

            for fut in futures:
                url, depth, allowed = futures[fut]
                data = fut.result()
                if not allowed:
                    data["robots_blocked"] = True
                html_z = data.pop("html_z", None)
                with STATE.lock:
                    STATE.pages[url] = data
                    if html_z and STATE.html_bytes < MAX_HTML_STORE_BYTES:
                        STATE.html_store[url] = html_z
                        STATE.html_bytes += len(html_z)
                    for res_url, res_type in data.get("resources", []):
                        if res_url not in STATE.resources:
                            STATE.resources[res_url] = {"type": res_type,
                                                        "status": None,
                                                        "sources": set()}
                        STATE.resources[res_url]["sources"].add(url)
                    for img_url, alt, lazy in data.get("images", []):
                        im = STATE.images.setdefault(img_url, {
                            "alts": set(), "missing_alt": False, "sources": set(),
                            "status": None, "kb": None, "ctype": "",
                            "lazy": 0, "uses": 0})
                        im["uses"] += 1
                        im["sources"].add(url)
                        if alt:
                            im["alts"].add(alt[:120])
                        else:
                            im["missing_alt"] = True
                        if lazy:
                            im["lazy"] += 1
                for link, anchor in data.get("links", []):
                    internal = same_domain(link, root_netloc)
                    with STATE.lock:
                        STATE.link_sources[link].add(url)
                        STATE.links.append({"source": url, "target": link,
                                            "anchor": anchor,
                                            "internal": internal})
                    if follow_links and internal and is_crawlable(link) \
                            and link not in queued and len(queued) < max_pages * 3:
                        queued.add(link)
                        frontier.append((link, depth + 1))
            with STATE.lock:
                STATE.queued_count = len(frontier)

    # ---------- Phase 2: resources + external links ----------
    with STATE.lock:
        stopped = STATE.stop_flag
        STATE.phase = "assets"
        STATE.queued_count = 0
        res_urls = list(STATE.resources.keys())[:MAX_RESOURCE_CHECKS]
        ext_urls = list({l["target"] for l in STATE.links
                         if not l["internal"]})[:MAX_EXTERNAL_CHECKS]
        STATE.assets_total = len(res_urls) + len(ext_urls)
        STATE.assets_done = 0

    if not stopped:
        with ThreadPoolExecutor(max_workers=workers) as pool:
            img_urls = [u for u in res_urls if STATE.resources[u]["type"] == "image"]
            other_urls = [u for u in res_urls if STATE.resources[u]["type"] != "image"]
            futs = {pool.submit(check_image, u): ("img", u) for u in img_urls}
            futs.update({pool.submit(check_status, u): ("res", u) for u in other_urls})
            futs.update({pool.submit(check_status, u): ("ext", u) for u in ext_urls})
            for fut in as_completed(futs):
                kind, u = futs[fut]
                result = fut.result()
                with STATE.lock:
                    if kind == "img":
                        STATE.resources[u]["status"] = result["status"]
                        if u in STATE.images:
                            STATE.images[u]["status"] = result["status"]
                            STATE.images[u]["kb"] = result["kb"]
                            STATE.images[u]["ctype"] = result["ctype"]
                    elif kind == "res":
                        STATE.resources[u]["status"] = result
                    else:
                        STATE.ext_status[u] = result
                    STATE.assets_done += 1
                    if STATE.stop_flag:
                        break

    with STATE.lock:
        STATE.phase = "done"
    try:
        snap = build_snapshot()
        changes = save_and_compare(snap)
    except Exception:
        changes = None
    with STATE.lock:
        STATE.changes = changes
        STATE.running = False


def _norm_issue(msg):
    return re.sub(r"\([^)]*\)", "", msg).strip()


def save_and_compare(snapshot):
    domain = urlparse(snapshot["start_url"]).netloc.replace(":", "_")
    if not domain:
        return None
    os.makedirs(HISTORY_DIR, exist_ok=True)
    path = os.path.join(HISTORY_DIR, f"{domain}.json")
    current = {
        "date": time.strftime("%Y-%m-%d %H:%M"),
        "issues": {r["url"]: sorted(i["msg"] for i in r["issues"])
                   for r in snapshot["rows"]},
        "statuses": {r["url"]: r["status"] for r in snapshot["rows"]},
    }
    changes = None
    if os.path.exists(path):
        try:
            prev = json.load(open(path, encoding="utf-8"))
        except Exception:
            prev = None
        if prev:
            prev_issues = {u: {_norm_issue(m) for m in ms}
                           for u, ms in prev.get("issues", {}).items()}
            cur_issues = {u: {_norm_issue(m) for m in ms}
                          for u, ms in current["issues"].items()}
            new_issues, resolved = [], []
            for u, ms in current["issues"].items():
                before = prev_issues.get(u, set())
                for m in ms:
                    if _norm_issue(m) not in before:
                        new_issues.append({"url": u, "msg": m})
            for u, ms in prev.get("issues", {}).items():
                now = cur_issues.get(u, set())
                if u not in current["statuses"]:
                    continue
                for m in ms:
                    if _norm_issue(m) not in now:
                        resolved.append({"url": u, "msg": m})
            prev_urls = set(prev.get("statuses", {}))
            cur_urls = set(current["statuses"])
            changes = {
                "prev_date": prev.get("date", "?"),
                "new_issues": new_issues,
                "resolved": resolved,
                "new_pages": sorted(cur_urls - prev_urls),
                "removed_pages": sorted(prev_urls - cur_urls),
            }
    else:
        changes = {"first_crawl": True}
    try:
        json.dump(current, open(path, "w", encoding="utf-8"))
    except Exception:
        pass
    return changes


def build_dashboard(rows, counts, redirects, res_rows, img_rows, link_rows):
    """Grade the crawl: what's good, what needs work, with suggestions."""
    pages_n = max(counts["total"], 1)
    html_ok = [r for r in rows if r["is_html"] and r["status"] == "200"]

    def urls_where(fn, data=None):
        out = []
        for x in (data or rows):
            if not fn(x):
                continue
            detail = ""
            if x.get("sources"):
                detail = "linked from: " + " | ".join(x["sources"][:3])
            out.append({"url": x["url"] if "url" in x else x.get("target", ""), "detail": detail})
        return out[:300]

    def has(r, cat, sub=""):
        sub = sub.lower()
        return any(i["cat"] == cat and sub in i["msg"].lower() for i in r["issues"])

    checks = [
        # (key, label, severity, suggestion, count, urls)
        ("broken", "Broken pages (4xx / 5xx / errors)", "high",
         "Restore the page or set a 301 redirect to the closest relevant page, then update the internal links pointing to it.",
         lambda r: has(r, "broken")),
        ("chains", "Redirect chains (2+ hops)", "medium",
         "Point the first URL directly at the final destination — each extra hop wastes crawl budget and link equity.",
         None),
        ("redirected_links", "Internal links pointing to redirects", "medium",
         "Update links on the 'Found On' pages to point directly to the final URL (see Response Codes → 3XX).",
         None),
        ("title_missing", "Missing page titles", "high",
         "Every indexable page needs a unique title. Write 30–60 characters with the primary keyword near the front.",
         lambda r: has(r, "title", "Title missing")),
        ("title_dup", "Duplicate page titles", "high",
         "Rewrite so each page targets a distinct search intent — duplicates make Google pick one page and ignore the rest.",
         lambda r: has(r, "title", "Duplicate")),
        ("title_len", "Titles too short / too long", "medium",
         "Keep titles 30–60 characters so they display fully in search results without truncation.",
         lambda r: has(r, "title", "too ")),
        ("desc_missing", "Missing meta descriptions", "medium",
         "Write 70–160 character descriptions that read like an ad for the page — they directly affect click-through rate.",
         lambda r: has(r, "desc", "Description missing")),
        ("desc_dup", "Duplicate meta descriptions", "medium",
         "Make each description specific to its page's content and intent.",
         lambda r: has(r, "desc", "Duplicate")),
        ("desc_len", "Descriptions too short / too long", "low",
         "Aim for 70–160 characters so the full description shows in search results.",
         lambda r: has(r, "desc", "too ")),
        ("h1_issues", "H1 problems (missing / multiple / duplicate)", "medium",
         "Give every page exactly one H1 that matches its topic; duplicates across pages confuse topical targeting.",
         lambda r: has(r, "h1")),
        ("canonical", "Canonical problems (missing / multiple)", "medium",
         "Add one self-referencing canonical tag per page (Rank Math does this automatically when configured).",
         lambda r: has(r, "canonical")),
        ("robots_conflict", "Sitemap vs robots.txt conflicts", "high",
         "Remove these URLs from the sitemap, or unblock them in robots.txt — sending Google contradictory signals hurts trust.",
         lambda r: has(r, "robots")),
        ("thin", "Thin content (under 300 words)", "medium",
         "Expand with genuinely useful content, or noindex pages that will always be thin (search pages, stubs).",
         lambda r: has(r, "thin", "Thin content")),
        ("dup_body", "Exact duplicate page content", "high",
         "Merge duplicate pages or canonicalise one to the other — identical content splits ranking signals.",
         lambda r: r["dup_body"]),
        ("noalt", "Images missing alt text", "medium",
         "Add descriptive alt text (naturally including keywords where relevant) — it's an image-search ranking factor and an accessibility requirement.",
         None),
        ("large_img", "Heavy images (over 200 KB)", "medium",
         "Compress and convert to WebP — usually the single biggest page-speed win on WordPress sites.",
         None),
        ("mixed", "Mixed content (http resources on https pages)", "high",
         "Change resource URLs to https:// — browsers block mixed content and it breaks the padlock.",
         lambda r: r["mixed"] > 0),
        ("http_pages", "Pages served over HTTP (not HTTPS)", "high",
         "Move to HTTPS with a 301 redirect — HTTPS is a confirmed ranking signal.",
         lambda r: r["is_html"] and r["status"] == "200" and not r["https"]),
        ("sec_headers", "Missing security headers", "low",
         "Add HSTS, X-Frame-Options, X-Content-Type-Options and CSP headers via .htaccess or a security plugin. Minor SEO impact, good practice.",
         lambda r: r["is_html"] and r["status"] == "200" and len(r.get("sec_missing", [])) > 0),
        ("slow", "Slow / heavy pages", "medium",
         "Investigate hosting response time, reduce plugins, enable caching, compress images.",
         lambda r: has(r, "speed")),
        ("schema_invalid", "Invalid structured data (JSON-LD)", "high",
         "Fix the JSON syntax and verify in Google's Rich Results Test — broken schema silently loses rich results.",
         lambda r: has(r, "schema", "Invalid")),
        ("no_schema", "Indexable pages without structured data", "low",
         "Add relevant schema (Article, FAQPage, BreadcrumbList) — an opportunity for rich results, not an error.",
         lambda r: r["is_html"] and r["status"] == "200" and r["indexable"] and not r["schema"]),
        ("lang", "Missing <html lang> attribute", "low",
         "Set the correct language code (en / gu / mr / hi) — helps search engines serve the right audience.",
         lambda r: has(r, "lang")),
        ("url_flags", "URL quality issues (uppercase, underscores, length)", "low",
         "Prefer short lowercase hyphenated URLs for new pages; don't rename existing ranking URLs just for this.",
         lambda r: len(r["url_flags"]) > 0),
        ("deep", "Deep pages (4+ clicks from home)", "low",
         "Link important deep pages from hub/category pages so crawlers and users reach them faster.",
         lambda r: has(r, "depth")),
        ("weak_inlinks", "Pages with 0–1 internal links (near-orphans)", "medium",
         "Add contextual internal links from related pages and hubs — pages with more inlinks rank better and get crawled more.",
         lambda r: r["is_html"] and r["status"] == "200" and r["indexable"] and r["inlinks"] <= 1 and r["depth"] > 0),
        ("broken_res", "Broken resources (images / CSS / JS)", "high",
         "Re-upload the missing file or fix the path — broken resources degrade rendering and user experience.",
         None),
        ("broken_ext", "Broken external links", "low",
         "Remove or replace dead outbound links. Note: 403s can be bot-blocking false alarms — verify manually.",
         None),
    ]

    def is_bad(s):
        return s == "ERROR" or (s.isdigit() and int(s) >= 400)

    def _src(lst):
        return " | ".join(lst[:3]) if lst else ""

    special = {
        "chains": [{"url": r["from"],
                    "detail": f"redirects to: {r['to']} ({r['hops']} hops)"
                              + (f" · linked from: {_src(r['sources'])}" if r.get("sources") else "")}
                   for r in redirects if r["hops"] > 1][:300],
        "redirected_links": [{"url": r["from"],
                              "detail": f"redirects to: {r['to']} · fix link on: {_src(r['sources'])}"}
                             for r in redirects if r.get("sources")][:300],
        "noalt": [{"url": r["url"], "detail": "used on: " + _src(r["sources"])}
                  for r in img_rows if r["missing_alt"]][:300],
        "large_img": [{"url": r["url"],
                       "detail": f"{r['kb']} KB · used on: {_src(r['sources'])}"}
                      for r in img_rows if r["kb"] is not None and r["kb"] > LARGE_IMG_KB][:300],
        "broken_res": [{"url": r["url"], "detail": "used on: " + _src(r["sources"])}
                       for r in res_rows if is_bad(r["status"])][:300],
        "broken_ext": [{"url": l["target"],
                        "detail": "linked from: " + " | ".join(
                            f"{e['source']} (anchor: \"{e['anchor']}\")" for e in l["examples"][:2])}
                       for l in link_rows if not l["internal"] and is_bad(l["status"])][:300],
    }

    weights = {"high": 30, "medium": 12, "low": 4}
    goods, works, penalty = [], [], 0.0
    good_labels = {
        "broken": "No broken pages — every URL responds correctly",
        "chains": "No redirect chains",
        "redirected_links": "No internal links pointing at redirects",
        "title_missing": "Every page has a title",
        "title_dup": "All page titles are unique",
        "title_len": "All titles are within 30–60 characters",
        "desc_missing": "Every page has a meta description",
        "desc_dup": "All meta descriptions are unique",
        "desc_len": "All descriptions are within 70–160 characters",
        "h1_issues": "H1s are clean — one unique H1 per page",
        "canonical": "Canonical tags are correctly set",
        "robots_conflict": "Sitemap and robots.txt agree",
        "thin": "No thin content — all pages exceed 300 words",
        "dup_body": "No exact duplicate content",
        "noalt": "All images have alt text",
        "large_img": "No oversized images (all under 200 KB)",
        "mixed": "No mixed content",
        "http_pages": "All pages served over HTTPS",
        "sec_headers": "Security headers present",
        "slow": "No slow or heavy pages",
        "schema_invalid": "All structured data is valid",
        "no_schema": "All indexable pages have structured data",
        "lang": "Language attribute set on all pages",
        "url_flags": "URLs are clean (lowercase, no underscores)",
        "deep": "No pages buried deeper than 3 clicks",
        "weak_inlinks": "All pages have 2+ internal links",
        "broken_res": "All resources (images / CSS / JS) load correctly",
        "broken_ext": "No broken external links",
    }
    for key, label, severity, suggestion, fn in checks:
        urls = special.get(key) if key in special else urls_where(fn)
        n = len(urls)
        if n == 0:
            goods.append(good_labels.get(key, label + " — OK"))
        else:
            works.append({"key": key, "label": label, "count": n,
                          "severity": severity, "suggestion": suggestion,
                          "urls": urls})
            penalty += weights[severity] * min(1.0, n / pages_n)
    score = max(0, round(100 - penalty))
    grade = ("A" if score >= 90 else "B" if score >= 75 else
             "C" if score >= 60 else "D" if score >= 40 else "E")
    sev_order = {"high": 0, "medium": 1, "low": 2}
    works.sort(key=lambda w: (sev_order[w["severity"]], -w["count"]))
    return {"score": score, "grade": grade, "goods": goods, "works": works}


def build_snapshot():
    with STATE.lock:
        pages = dict(STATE.pages)
        redirects = [dict(r) for r in STATE.redirects]
        link_sources = {k: sorted(v) for k, v in STATE.link_sources.items()}
        link_source_counts = {k: len(v) for k, v in STATE.link_sources.items()}
        links = list(STATE.links)
        resources = {u: {"type": d["type"], "status": d["status"],
                         "sources": sorted(d["sources"])[:5],
                         "inlinks": len(d["sources"])}
                     for u, d in STATE.resources.items()}
        images = {u: {"alts": sorted(d["alts"])[:3], "missing_alt": d["missing_alt"],
                      "sources": sorted(d["sources"])[:5], "status": d["status"],
                      "kb": d["kb"], "ctype": d["ctype"],
                      "lazy": d["lazy"], "uses": d["uses"]}
                  for u, d in STATE.images.items()}
        ext_status = dict(STATE.ext_status)
        running = STATE.running
        phase = STATE.phase
        mode = STATE.mode
        robots_found = STATE.robots_found
        changes = STATE.changes
        queued = STATE.queued_count
        assets_total = STATE.assets_total
        assets_done = STATE.assets_done
        start_url = STATE.start_url
        error_note = STATE.error_note

    root_netloc = urlparse(start_url).netloc
    for r in redirects:
        r["sources"] = link_sources.get(r["from"], [])[:10]
    redirect_map = {r["from"]: r for r in redirects}

    def indexability(u, d):
        status = d.get("status")
        if status == "BLOCKED":
            return False, "Blocked by robots.txt"
        if status == "ERROR":
            return False, "Fetch error"
        if u in redirect_map:
            return False, f"Redirect ({redirect_map[u]['first_status']})"
        if isinstance(status, int) and status >= 400:
            return False, f"HTTP {status}"
        if isinstance(status, int) and 300 <= status < 400:
            return False, f"Redirect ({status})"
        if d.get("robots_blocked"):
            return False, "Blocked by robots.txt (but fetched via sitemap)"
        if d.get("noindex"):
            return False, "Noindex"
        canon = d.get("canonical", "")
        if canon and canon.rstrip("/") != d.get("final_url", u).rstrip("/"):
            return False, "Canonicalised to another URL"
        if "title" not in d:
            return False, "Non-HTML"
        return True, "Indexable"

    html_pages = {u: d for u, d in pages.items() if indexability(u, d)[0]}
    titles, descs, h1s, h2s, kws, bodies = (defaultdict(list) for _ in range(6))
    for u, d in html_pages.items():
        if d.get("title"): titles[d["title"]].append(u)
        if d.get("description"): descs[d["description"]].append(u)
        if d.get("h1"): h1s[d["h1"]].append(u)
        if d.get("h2"): h2s[d["h2"]].append(u)
        if d.get("keywords"): kws[d["keywords"]].append(u)
        if d.get("body_hash"): bodies[d["body_hash"]].append(u)
    dup_titles = {t for t, us in titles.items() if len(us) > 1}
    dup_descs = {t for t, us in descs.items() if len(us) > 1}
    dup_h1s = {t for t, us in h1s.items() if len(us) > 1}
    dup_h2s = {t for t, us in h2s.items() if len(us) > 1}
    dup_kws = {t for t, us in kws.items() if len(us) > 1}
    dup_bodies = {t: len(us) for t, us in bodies.items() if len(us) > 1}

    rows = []
    for u, d in sorted(pages.items()):
        issues = []
        status = d.get("status")
        is_html = "title" in d
        indexable, index_reason = indexability(u, d)

        if u in redirect_map:
            r = redirect_map[u]
            issues.append(("redirect",
                           f"{r['first_status']} redirect ({r['hops']} hop{'s' if r['hops'] > 1 else ''})"))
        if status == "BLOCKED":
            issues.append(("robots", "Blocked by robots.txt"
                           + (" — but in sitemap!" if mode == "sitemap" else "")))
        elif d.get("robots_blocked"):
            issues.append(("robots", "In sitemap but blocked by robots.txt"))
        if status == "ERROR":
            issues.append(("broken", "Fetch error: " + d.get("error", "")))
        elif isinstance(status, int) and status >= 400:
            issues.append(("broken", f"HTTP {status}"))

        if indexable and is_html and status == 200:
            t = d.get("title", "")
            if not t: issues.append(("title", "Title missing"))
            elif len(t) < TITLE_MIN: issues.append(("title", f"Title too short ({len(t)})"))
            elif len(t) > TITLE_MAX: issues.append(("title", f"Title too long ({len(t)})"))
            if t and t in dup_titles: issues.append(("title", "Duplicate title"))
            desc = d.get("description", "")
            if not desc: issues.append(("desc", "Description missing"))
            elif len(desc) < DESC_MIN: issues.append(("desc", f"Description too short ({len(desc)})"))
            elif len(desc) > DESC_MAX: issues.append(("desc", f"Description too long ({len(desc)})"))
            if desc and desc in dup_descs: issues.append(("desc", "Duplicate description"))
            hc = d.get("h1_count", 0)
            if hc == 0: issues.append(("h1", "H1 missing"))
            elif hc > 1: issues.append(("h1", f"Multiple H1s ({hc})"))
            if d.get("h1") and d["h1"] in dup_h1s: issues.append(("h1", "Duplicate H1"))
            if not d.get("canonical", ""):
                issues.append(("canonical", "Canonical missing"))
            if d.get("canonical_count", 0) > 1:
                issues.append(("canonical", f"Multiple canonicals ({d['canonical_count']})"))
            if d.get("imgs_no_alt", 0) > 0:
                issues.append(("img", f"{d['imgs_no_alt']} images missing alt"))
            wc = d.get("word_count", 0)
            if wc < THIN_WORDS:
                issues.append(("thin", f"Thin content ({wc} words)"))
            if d.get("body_hash") in dup_bodies:
                issues.append(("thin", f"Exact duplicate body ({dup_bodies[d['body_hash']]} pages)"))
            if mode == "spider" and d.get("depth", 0) >= DEEP_CLICKS:
                issues.append(("depth", f"Deep page ({d['depth']} clicks from home)"))
            if d.get("ms", 0) > SLOW_MS:
                issues.append(("speed", f"Slow response ({d['ms']/1000:.1f}s)"))
            if d.get("kb", 0) > LARGE_KB:
                issues.append(("speed", f"Large HTML ({d['kb']:.0f} KB)"))
            if len(d.get("resources", [])) > MANY_RESOURCES:
                issues.append(("speed", f"Many resources ({len(d['resources'])})"))
            if d.get("schema_invalid", 0) > 0:
                issues.append(("schema", f"Invalid JSON-LD ({d['schema_invalid']} block/s)"))
            if not d.get("lang"):
                issues.append(("lang", "Missing <html lang> attribute"))
            if d.get("mixed", 0) > 0:
                issues.append(("security", f"Mixed content ({d['mixed']} http resources)"))

        # URL quality flags
        path = urlparse(u).path
        url_flags = []
        if any(ord(ch) > 127 for ch in u): url_flags.append("non-ascii")
        if "_" in path: url_flags.append("underscores")
        if any(c.isupper() for c in path): url_flags.append("uppercase")
        if "?" in u: url_flags.append("parameters")
        if len(u) > URL_MAX_LEN: url_flags.append(f"over {URL_MAX_LEN} chars")

        rows.append({
            "url": u, "status": str(status), "depth": d.get("depth", 0),
            "title": d.get("title", ""), "h1": d.get("h1", ""),
            "description": d.get("description", ""),
            "keywords": d.get("keywords", ""),
            "kw_dup": bool(d.get("keywords")) and d.get("keywords") in dup_kws,
            "h1_all": d.get("h1_all", ""), "h1_count": d.get("h1_count", 0),
            "h2": d.get("h2", ""), "h2_all": d.get("h2_all", ""),
            "h2_count": d.get("h2_count", 0),
            "h2_dup": bool(d.get("h2")) and d.get("h2") in dup_h2s,
            "canonical": d.get("canonical", ""),
            "canonical_count": d.get("canonical_count", 0),
            "words": d.get("word_count", ""),
            "dup_body": d.get("body_hash", "") in dup_bodies,
            "dup_body_n": dup_bodies.get(d.get("body_hash", ""), 0),
            "ms": d.get("ms", ""), "kb": d.get("kb", ""),
            "res_count": len(d.get("resources", [])) if is_html else "",
            "lang": d.get("lang", ""),
            "schema": sorted(set(d.get("schema_types", []))),
            "https": d.get("https", u.startswith("https://")),
            "mixed": d.get("mixed", 0),
            "sec_missing": d.get("sec_missing", []),
            "url_flags": url_flags,
            "is_html": is_html,
            "inlinks": link_source_counts.get(u, 0),
            "indexable": indexable, "index_reason": index_reason,
            "issues": [{"cat": c, "msg": m} for c, m in issues],
            "sources": link_sources.get(u, [])[:5]
                if (status == "ERROR" or (isinstance(status, int) and status >= 400)) else [],
        })

    # resources / images rows
    res_rows = []
    for u, d in sorted(resources.items()):
        res_rows.append({"url": u, "type": d["type"],
                         "status": "..." if d["status"] is None else str(d["status"]),
                         "sources": d["sources"], "inlinks": d["inlinks"],
                         "internal": same_domain(u, root_netloc) if root_netloc else True})

    def img_format(u, ctype):
        if ctype and "/" in ctype:
            f = ctype.split("/")[-1].split(";")[0].lower().replace("jpeg", "jpg")
            if f not in ("octet-stream", "html", "plain"):
                return f
        p = urlparse(u).path.lower()
        for ext in (".webp", ".avif", ".svg", ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".ico"):
            if p.endswith(ext):
                return ext[1:].replace("jpeg", "jpg")
        return "?"

    img_rows = []
    for u, d in sorted(images.items()):
        img_rows.append({
            "url": u,
            "status": "..." if d["status"] is None else str(d["status"]),
            "kb": d["kb"], "format": img_format(u, d["ctype"]),
            "missing_alt": d["missing_alt"], "alts": d["alts"],
            "lazy": d["lazy"], "uses": d["uses"], "sources": d["sources"]})

    # links aggregated by target
    agg = {}
    for l in links:
        key = l["target"]
        if key not in agg:
            agg[key] = {"target": key, "internal": l["internal"],
                        "count": 0, "examples": []}
        agg[key]["count"] += 1
        if len(agg[key]["examples"]) < 3:
            agg[key]["examples"].append({"source": l["source"], "anchor": l["anchor"]})
    link_rows = []
    for key, a in sorted(agg.items()):
        if a["internal"]:
            pd = pages.get(key)
            st = str(pd["status"]) if pd else "not crawled"
        else:
            st = ext_status.get(key)
            st = "..." if st is None else str(st)
        a["status"] = st
        link_rows.append(a)

    def is_bad(s):
        return s == "ERROR" or (s.isdigit() and int(s) >= 400)

    counts = {
        "total": len(rows),
        "internal": len(rows) + sum(1 for r in res_rows if r["internal"]),
        "external": sum(1 for l in link_rows if not l["internal"]),
        "security": sum(1 for r in rows if r["is_html"] and (not r["https"] or r["mixed"] or r["sec_missing"])),
        "response": len(rows),
        "url": sum(1 for r in rows if r["url_flags"]),
        "title": sum(1 for r in rows if any(i["cat"] == "title" for i in r["issues"])),
        "desc": sum(1 for r in rows if any(i["cat"] == "desc" for i in r["issues"])),
        "keywords": sum(1 for r in rows if r["keywords"]),
        "h1": sum(1 for r in rows if any(i["cat"] == "h1" for i in r["issues"])),
        "h2": sum(1 for r in rows if r["is_html"] and r["status"] == "200" and (r["h2_count"] == 0 or r["h2_dup"])),
        "content": sum(1 for r in rows if any(i["cat"] == "thin" for i in r["issues"])),
        "images": len(img_rows),
        "images_noalt": sum(1 for r in img_rows if r["missing_alt"]),
        "canonical": sum(1 for r in rows if any(i["cat"] == "canonical" for i in r["issues"])
                         or (not r["indexable"] and "Canonicalised" in r["index_reason"])),
        "speed": sum(1 for r in rows if any(i["cat"] == "speed" for i in r["issues"])),
        "changes": (len(changes.get("new_issues", [])) if changes and not changes.get("first_crawl") else 0),
        "resources_broken": sum(1 for r in res_rows if is_bad(r["status"])),
        "links_broken": sum(1 for l in link_rows if is_bad(l["status"])),
    }
    dashboard = build_dashboard(rows, counts, redirects, res_rows, img_rows, link_rows)
    return {"running": running, "phase": phase, "queued": queued,
            "mode": mode, "robots_found": robots_found,
            "assets_total": assets_total, "assets_done": assets_done,
            "start_url": start_url, "error_note": error_note,
            "changes": changes, "dashboard": dashboard,
            "rows": rows, "redirects": redirects,
            "resources": res_rows, "images": img_rows,
            "links": link_rows, "counts": counts}


_STOPWORDS = set("""a an and are as at be by for from has have how in is it its of on or that the
this to was were will with what when where which who your you our we they i not no all can if
do does did been about into more most other some such than then there these those out up so""".split())


def _page_text(url):
    """Decompressed, tag-stripped, lowercased text of a crawled page."""
    z = STATE.html_store.get(url)
    if z is None:
        return ""
    try:
        html = zlib.decompress(z).decode("utf-8", "ignore")
    except Exception:
        return ""
    html = re.sub(r"<script[^>]*>.*?</script>|<style[^>]*>.*?</style>", " ", html,
                  flags=re.S | re.I)
    text = re.sub(r"<[^>]+>", " ", html)
    return re.sub(r"\s+", " ", text).lower()


def _primary_phrase(d):
    """The phrase that identifies a page's topic: H1, else title before separator."""
    for cand in (d.get("h1", ""), re.split(r"[|\-–—]", d.get("title", ""))[0]):
        cand = cand.strip().lower()
        if len(cand) >= 8 and len(cand.split()) >= 2:
            return cand
    return ""


def _top_terms(text, n=15):
    freq = {}
    for w in re.findall(r"\w{4,}", text, re.UNICODE):
        if w in _STOPWORDS or w.isdigit():
            continue
        freq[w] = freq.get(w, 0) + 1
    return set(sorted(freq, key=freq.get, reverse=True)[:n])


def link_suggestions():
    """Suggest internal links: unlinked mentions + topic-overlap matches."""
    with STATE.lock:
        pages = dict(STATE.pages)
        links = list(STATE.links)

    cand = {}
    for u, d in pages.items():
        if d.get("status") == 200 and "title" in d and not d.get("noindex"):
            canon = d.get("canonical", "")
            if canon and canon.rstrip("/") != d.get("final_url", u).rstrip("/"):
                continue
            cand[u] = d
    cand_urls = list(cand.keys())[:2000]   # O(n²) guard

    existing = defaultdict(set)   # source -> targets
    inlink_count = defaultdict(int)
    for l in links:
        existing[l["source"]].add(l["target"])
        if l["internal"]:
            inlink_count[l["target"]] += 1

    texts = {u: _page_text(u) for u in cand_urls}
    raw_terms = {u: _top_terms(texts[u], 25) for u in cand_urls}
    # boilerplate filter: terms appearing on most pages (nav/footer/filler) carry no topical signal
    df = defaultdict(int)
    for ts in raw_terms.values():
        for w in ts:
            df[w] += 1
    n_pages = max(len(cand_urls), 1)
    boiler = {w for w, c in df.items() if n_pages >= 5 and c > 0.6 * n_pages}
    terms = {u: set(list(ts - boiler)[:15]) for u, ts in raw_terms.items()}
    phrases = {u: _primary_phrase(cand[u]) for u in cand_urls}

    targets = sorted(cand_urls, key=lambda u: inlink_count.get(u, 0))[:100]
    out = []
    for t in targets:
        t_phrase = phrases[t]
        sugg = []
        for s in cand_urls:
            if s == t or t in existing.get(s, set()):
                continue
            if t_phrase and t_phrase in texts.get(s, ""):
                sugg.append({"source": s, "type": "mention", "anchor": t_phrase,
                             "evidence": f'page text contains "{t_phrase}" with no link'})
                continue
            shared = terms[t] & terms.get(s, set())
            if len(shared) >= 4:
                sugg.append({"source": s, "type": "topic",
                             "anchor": t_phrase or cand[t].get("title", t),
                             "evidence": "shared topics: " + ", ".join(sorted(shared)[:5])})
        if sugg:
            sugg.sort(key=lambda x: 0 if x["type"] == "mention" else 1)
            out.append({"target": t, "title": cand[t].get("title", ""),
                        "inlinks": inlink_count.get(t, 0),
                        "suggestions": sugg[:5]})
    mentions_total = sum(1 for o in out for s in o["suggestions"] if s["type"] == "mention")
    return {"targets": out, "mentions": mentions_total,
            "pages_analyzed": len(cand_urls)}


def custom_search(q, mode):
    """Search stored page HTML. mode: contains | not"""
    q_low = q.lower()
    with STATE.lock:
        store = dict(STATE.html_store)
        page_urls = [u for u, d in STATE.pages.items()
                     if d.get("status") == 200 and "title" in d]
    results = []
    for u in sorted(page_urls):
        z = store.get(u)
        if z is None:
            continue
        try:
            html = zlib.decompress(z).decode("utf-8", "ignore")
        except Exception:
            continue
        low = html.lower()
        n = low.count(q_low)
        if mode == "contains" and n > 0:
            i = low.find(q_low)
            snippet = re.sub(r"<[^>]+>", " ", html[max(0, i - 80):i + len(q) + 80])
            snippet = re.sub(r"\s+", " ", snippet).strip()
            results.append({"url": u, "count": n, "snippet": snippet[:220]})
        elif mode == "not" and n == 0:
            results.append({"url": u, "count": 0, "snippet": ""})
    return results


def url_detail(url):
    """Everything known about one URL: inlinks, outlinks, images on page."""
    with STATE.lock:
        links = list(STATE.links)
        pages = dict(STATE.pages)
        ext_status = dict(STATE.ext_status)
        images = {u: {"alts": sorted(d["alts"])[:3], "missing_alt": d["missing_alt"],
                      "sources": set(d["sources"]), "kb": d["kb"],
                      "status": d["status"]}
                  for u, d in STATE.images.items()}
        resources = {u: {"type": d["type"], "status": d["status"],
                         "sources": set(d["sources"])}
                     for u, d in STATE.resources.items()}

    inlinks = [{"source": l["source"], "anchor": l["anchor"]}
               for l in links if l["target"] == url][:200]

    outlinks = []
    for l in links:
        if l["source"] != url:
            continue
        if l["internal"]:
            pd = pages.get(l["target"])
            st = str(pd["status"]) if pd else "not crawled"
        else:
            st = ext_status.get(l["target"])
            st = "?" if st is None else str(st)
        outlinks.append({"target": l["target"], "anchor": l["anchor"],
                         "internal": l["internal"], "status": st})
        if len(outlinks) >= 300:
            break

    page_images = []
    for iu, d in images.items():
        if url in d["sources"]:
            page_images.append({"url": iu, "alts": list(d["alts"]),
                                "missing_alt": d["missing_alt"],
                                "kb": d["kb"],
                                "status": "?" if d["status"] is None else str(d["status"])})
        if len(page_images) >= 100:
            break

    page_resources = []
    for ru, d in resources.items():
        if url in d["sources"] and d["type"] != "image":
            page_resources.append({"url": ru, "type": d["type"],
                                   "status": "?" if d["status"] is None else str(d["status"])})
        if len(page_resources) >= 100:
            break

    return {"url": url, "inlinks": inlinks, "outlinks": outlinks,
            "images": page_images, "resources": page_resources}


HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Makdi — SEO Crawler</title>
<style>
  * { margin:0; padding:0; box-sizing:border-box; }
  body { font-family:'Segoe UI',system-ui,sans-serif; background:#f0f4f8; color:#1a2b3c; padding:20px; }
  .wrap { max-width:1340px; margin:0 auto; }
  h1 { font-size:1.45rem; color:#1F4E79; margin-bottom:14px; }
  .inputbar { display:flex; gap:10px; flex-wrap:wrap; margin-bottom:16px; }
  select, input[type=text], input[type=number] { padding:11px; border:2px solid #cdd9e5; border-radius:8px; font-size:.9rem; }
  select { font-weight:600; color:#1F4E79; }
  input[type=text] { flex:1; min-width:280px; }
  input[type=number] { width:110px; }
  input:focus, select:focus { outline:none; border-color:#1F4E79; }
  button { background:#1F4E79; color:#fff; border:none; padding:11px 24px; border-radius:8px;
    font-size:.95rem; font-weight:700; cursor:pointer; }
  button:hover { background:#163a5c; }
  button:disabled { background:#9db2c4; cursor:not-allowed; }
  button.stop { background:#a11212; }
  .progress { display:none; align-items:center; gap:12px; margin-bottom:14px; font-size:.9rem; color:#5a6b7c; }
  .spinner { width:16px; height:16px; border:3px solid #cdd9e5; border-top-color:#1F4E79;
    border-radius:50%; animation:spin .8s linear infinite; }
  @keyframes spin { to { transform:rotate(360deg); } }
  .tabs { display:flex; gap:5px; flex-wrap:wrap; margin-bottom:12px; }
  .tab { background:#fff; border:2px solid #cdd9e5; padding:6px 12px; border-radius:20px;
    font-size:.8rem; font-weight:600; cursor:pointer; color:#3a4b5c; }
  .tab.active { background:#1F4E79; border-color:#1F4E79; color:#fff; }
  .tab .n { background:rgba(0,0,0,.12); border-radius:10px; padding:1px 7px; margin-left:5px; font-size:.74rem; }
  .tab.active .n { background:rgba(255,255,255,.25); }
  .tab .nb { background:#ffd6d6; color:#a11212; border-radius:10px; padding:1px 7px; margin-left:4px; font-size:.74rem; font-weight:700; }
  .subbar { display:flex; gap:8px; margin-bottom:10px; flex-wrap:wrap; }
  .sub { background:#fff; border:2px solid #cdd9e5; padding:5px 14px; border-radius:16px;
    font-size:.78rem; font-weight:600; cursor:pointer; color:#3a4b5c; }
  .sub.active { background:#3a4b5c; border-color:#3a4b5c; color:#fff; }
  table { width:100%; border-collapse:collapse; background:#fff; border-radius:10px;
    overflow:hidden; box-shadow:0 2px 8px rgba(0,0,0,.06); }
  th { background:#1F4E79; color:#fff; text-align:left; padding:9px 12px; font-size:.76rem;
    text-transform:uppercase; letter-spacing:.4px; }
  td { padding:9px 12px; border-bottom:1px solid #edf2f7; font-size:.84rem; vertical-align:top; word-break:break-word; }
  .badge { display:inline-block; padding:2px 9px; border-radius:11px; font-weight:700; font-size:.75rem; }
  .ok { background:#d1f5dd; color:#0f6d33; }
  .warn { background:#fff2c4; color:#8a6400; }
  .bad { background:#ffd6d6; color:#a11212; }
  .gray { background:#e4eaf0; color:#5a6b7c; }
  .type { display:inline-block; padding:2px 9px; border-radius:11px; font-weight:700; font-size:.73rem;
    background:#eef4fb; color:#1F4E79; }
  .issue { display:inline-block; background:#fef0f0; color:#a11212; border:1px solid #f3c6c6;
    padding:1px 8px; border-radius:10px; font-size:.73rem; margin:1px 3px 1px 0; font-weight:600; }
  .issue.title, .issue.desc, .issue.redirect { background:#fff8e1; color:#8a6400; border-color:#ecd9a0; }
  .issue.h1, .issue.canonical, .issue.img, .issue.lang { background:#eef4fb; color:#1F4E79; border-color:#c3d6ea; }
  .issue.thin, .issue.depth { background:#f3ecfb; color:#5b2d91; border-color:#d5c3ee; }
  .issue.robots, .issue.schema, .issue.security { background:#fef0f0; color:#a11212; border-color:#f3c6c6; }
  .issue.speed { background:#fff0e6; color:#b3541e; border-color:#f0cbb0; }
  .meta { font-size:.72rem; color:#8a99a8; margin-top:2px; }
  .src { font-size:.76rem; color:#8a99a8; margin-top:3px; }
  .anchor { color:#3a4b5c; font-style:italic; }
  .chain { font-size:.76rem; color:#5a6b7c; font-family:Consolas,monospace; }
  .chain b { color:#8a6400; }
  .empty { text-align:center; color:#8a99a8; padding:40px 0; background:#fff; border-radius:10px; }
  .urlcell a { color:#1F4E79; text-decoration:none; }
  .urlcell a:hover { text-decoration:underline; }
  .chgnew { color:#a11212; font-weight:700; }
  .chgres { color:#0f6d33; font-weight:700; }
  .chghead { font-size:.95rem; font-weight:700; color:#1F4E79; margin:16px 0 8px; }
  .searchbar { display:flex; gap:8px; margin-bottom:12px; flex-wrap:wrap; }
  .searchbar input[type=text] { flex:1; min-width:200px; }
  /* ---- dashboard ---- */
  .dash-top { display:flex; gap:24px; align-items:center; background:#fff; border-radius:12px;
    padding:22px 26px; box-shadow:0 2px 8px rgba(0,0,0,.06); margin-bottom:16px; flex-wrap:wrap; }
  .dash-score { width:120px; height:120px; border-radius:50%; border:6px solid;
    display:flex; flex-direction:column; align-items:center; justify-content:center; flex:0 0 auto; }
  .dash-num { font-size:2.3rem; font-weight:800; line-height:1; }
  .dash-grade { font-size:.8rem; color:#5a6b7c; font-weight:700; margin-top:4px; }
  .dash-site { font-size:1.15rem; font-weight:700; color:#1F4E79; word-break:break-all; }
  .dash-btns { display:flex; gap:10px; margin-top:12px; flex-wrap:wrap; }
  .dash-cols { display:grid; grid-template-columns:1fr 1.4fr; gap:16px; }
  @media (max-width:900px) { .dash-cols { grid-template-columns:1fr; } }
  .dash-col { background:#fff; border-radius:12px; padding:18px 20px; box-shadow:0 2px 8px rgba(0,0,0,.06); }
  .dash-good { padding:7px 10px; border-bottom:1px solid #edf2f7; font-size:.88rem; color:#0f6d33; }
  .dash-good:last-child { border-bottom:none; }
  .dash-work { border:1px solid #e2e9f0; border-radius:10px; padding:12px 14px; margin-bottom:10px; }
  .dash-work-head { display:flex; gap:8px; align-items:center; cursor:pointer; flex-wrap:wrap; font-size:.9rem; }
  .dash-expand { margin-left:auto; color:#8a99a8; font-size:.9rem; }
  .dash-sugg { font-size:.84rem; color:#5a6b7c; margin-top:6px; }
  .dash-urls { margin-top:8px; border-top:1px dashed #e2e9f0; padding-top:8px; max-height:260px; overflow-y:auto; }
  /* ---- charts ---- */
  .chartrow { display:grid; grid-template-columns:1fr 1fr; gap:16px; margin-bottom:14px; }
  @media (max-width:820px) { .chartrow { grid-template-columns:1fr; } }
  .chartcard { background:#fff; border-radius:12px; padding:14px 18px 8px; box-shadow:0 2px 8px rgba(0,0,0,.06); }
  .charttitle { font-size:.84rem; font-weight:700; color:#1F4E79; margin-bottom:4px; }
  .legend { display:flex; gap:12px; flex-wrap:wrap; margin-top:4px; }
  .legend span { font-size:.76rem; color:#5a6b7c; display:flex; align-items:center; gap:5px; }
  .legend i { width:10px; height:10px; border-radius:3px; display:inline-block; }
  /* ---- URL detail panel ---- */
  tr[data-url] { cursor:pointer; }
  tr[data-url]:hover td { background:#f6f9fc; }
  .overlay { display:none; position:fixed; inset:0; background:rgba(15,35,60,.35); z-index:40; }
  .detail { display:none; position:fixed; top:0; right:0; bottom:0; width:min(560px, 94vw);
    background:#fff; z-index:50; box-shadow:-8px 0 30px rgba(15,35,60,.25);
    overflow-y:auto; padding:20px 22px; }
  .detail-close { position:sticky; top:0; float:right; background:#eef4fb; border:none;
    color:#1F4E79; font-size:1rem; font-weight:700; border-radius:8px; padding:6px 13px; cursor:pointer; }
  .detail h2 { font-size:1rem; color:#1F4E79; word-break:break-all; margin:4px 0 10px; padding-right:70px; }
  .detail-sec { font-size:.8rem; font-weight:800; color:#1F4E79; text-transform:uppercase;
    letter-spacing:.06em; margin:18px 0 8px; border-bottom:2px solid #eef4fb; padding-bottom:4px; }
  .ov { width:100%; border-collapse:collapse; }
  .ov td { padding:5px 8px; border-bottom:1px solid #f2f6fa; font-size:.83rem; vertical-align:top; }
  .ov td:first-child { color:#8a99a8; width:130px; white-space:nowrap; }
  .dl-item { padding:7px 4px; border-bottom:1px solid #f2f6fa; font-size:.82rem; word-break:break-all; }
  .dl-item a { color:#1F4E79; text-decoration:none; }
  .dl-item a:hover { text-decoration:underline; }
  @media print { .overlay, .detail { display:none !important; } }
  @media print {
    body { background:#fff; padding:0; }
    .inputbar, .tabs, .progress, .dash-btns { display:none !important; }
    .dash-top, .dash-col, .dash-work, .chartcard { box-shadow:none; border:1px solid #ccc; }
    .dash-cols { grid-template-columns:1fr; }
    .chartrow { grid-template-columns:1fr 1fr; }
    .dash-urls { display:block !important; max-height:none; overflow:visible; }
    .dash-expand { display:none; }
  }
</style>
</head>
<body>
<div class="wrap">
  <h1 style="display:flex;align-items:center;gap:10px">
    <svg viewBox="0 0 48 48" style="width:30px;height:30px" aria-hidden="true">
      <g stroke="#E8A33D" stroke-width="2.6" stroke-linecap="round" fill="none">
        <path d="M14 12 L20 19 M34 12 L28 19 M8 24 L18 24 M40 24 L30 24 M14 38 L20 30 M34 38 L28 30"/>
      </g>
      <circle cx="24" cy="21" r="4.5" fill="#E8A33D"/>
      <ellipse cx="24" cy="30" rx="6" ry="7" fill="#E8A33D"/>
      <circle cx="22" cy="20" r="1" fill="#1F4E79"/><circle cx="26" cy="20" r="1" fill="#1F4E79"/>
    </svg>
    MAKDI <span style="font-size:.8rem;color:#8a99a8;font-weight:400;letter-spacing:0">v""" + VERSION + r""" — the free unlimited SEO crawler</span></h1>

  <div class="inputbar">
    <select id="mode">
      <option value="spider">🕷 Crawl domain</option>
      <option value="sitemap">🗺 Crawl sitemap URLs</option>
    </select>
    <input type="text" id="domain" placeholder="Enter domain (e.g. example.com) or sitemap URL (.xml)" onkeydown="if(event.key==='Enter')startCrawl()">
    <input type="number" id="maxpages" value="500" min="10" max="10000" title="Max pages">
    <button id="startBtn" onclick="startCrawl()">Start Crawl</button>
    <button id="stopBtn" class="stop" onclick="stopCrawl()" style="display:none">Stop</button>
  </div>

  <div class="progress" id="progress">
    <div class="spinner"></div>
    <span id="progressText">Crawling...</span>
  </div>

  <div class="tabs" id="tabs" style="display:none">
    <div class="tab active" data-f="dashboard" onclick="setFilter('dashboard')">📊 Dashboard</div>
    <div class="tab" data-f="internal" onclick="setFilter('internal')">Internal<span class="n" id="n-internal">0</span></div>
    <div class="tab" data-f="external" onclick="setFilter('external')">External<span class="n" id="n-external">0</span><span class="nb" id="nb-external" style="display:none">0</span></div>
    <div class="tab" data-f="security" onclick="setFilter('security')">Security<span class="n" id="n-security">0</span></div>
    <div class="tab" data-f="response" onclick="setFilter('response')">Response Codes<span class="n" id="n-response">0</span></div>
    <div class="tab" data-f="url" onclick="setFilter('url')">URL<span class="n" id="n-url">0</span></div>
    <div class="tab" data-f="title" onclick="setFilter('title')">Page Titles<span class="n" id="n-title">0</span></div>
    <div class="tab" data-f="desc" onclick="setFilter('desc')">Meta Description<span class="n" id="n-desc">0</span></div>
    <div class="tab" data-f="keywords" onclick="setFilter('keywords')">Meta Keywords<span class="n" id="n-keywords">0</span></div>
    <div class="tab" data-f="h1" onclick="setFilter('h1')">H1<span class="n" id="n-h1">0</span></div>
    <div class="tab" data-f="h2" onclick="setFilter('h2')">H2<span class="n" id="n-h2">0</span></div>
    <div class="tab" data-f="content" onclick="setFilter('content')">Content<span class="n" id="n-content">0</span></div>
    <div class="tab" data-f="images" onclick="setFilter('images')">Images<span class="n" id="n-images">0</span><span class="nb" id="nb-images" style="display:none">0</span></div>
    <div class="tab" data-f="canonical" onclick="setFilter('canonical')">Canonicals<span class="n" id="n-canonical">0</span></div>
    <div class="tab" data-f="speed" onclick="setFilter('speed')">PageSpeed<span class="n" id="n-speed">0</span></div>
    <div class="tab" data-f="search" onclick="setFilter('search')">Custom Search</div>
    <div class="tab" data-f="linksugg" onclick="setFilter('linksugg')">Link Suggestions</div>
    <div class="tab" data-f="changes" onclick="setFilter('changes')">Changes<span class="n" id="n-changes">0</span></div>
    <div class="tab" style="margin-left:auto;background:#eef4fb;border-color:#c3d6ea" onclick="copyUrls()" title="Copy URLs of current view">📋 Copy URLs</div>
    <div class="tab" style="background:#eef4fb;border-color:#c3d6ea" onclick="exportCsv()" title="Download current view as CSV">⬇ Export CSV</div>
  </div>

  <div id="tablebox"></div>
</div>

<div class="overlay" id="overlay" onclick="closeDetails()"></div>
<div class="detail" id="detail"></div>

<script>
let snapshot = null, filter = 'dashboard', pollTimer = null;
let tabSub = {};
let searchResults = null, searchQ = '', searchMode = 'contains';
let linkSugg = null, linkSuggLoading = false, linkSuggFilter = 'all';

function esc(s) {
  return String(s??'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}
function isBad(s) { return s === 'ERROR' || (parseInt(s) >= 400); }
function hasMsg(r, cat, substr) {
  substr = substr.toLowerCase();
  return r.issues.some(i => i.cat === cat && i.msg.toLowerCase().includes(substr));
}
function htmlRows() { return snapshot.rows.filter(r => r.is_html && r.status === '200'); }

/* ---------------- sub-filter definitions ---------------- */
const SUBDEFS = {
  security: [
    ['all','ALL PAGES', r => r.is_html && r.status === '200'],
    ['http','HTTP (NOT HTTPS)', r => r.is_html && r.status === '200' && !r.https],
    ['mixed','MIXED CONTENT', r => r.mixed > 0],
    ['hsts','MISSING HSTS', r => r.is_html && r.sec_missing.includes('HSTS')],
    ['xfo','MISSING X-FRAME-OPTIONS', r => r.is_html && r.sec_missing.includes('X-Frame-Options')],
    ['xcto','MISSING X-CONTENT-TYPE', r => r.is_html && r.sec_missing.includes('X-Content-Type-Options')],
    ['csp','MISSING CSP', r => r.is_html && r.sec_missing.includes('CSP')],
  ],
  response: [
    ['all','ALL', r => true],
    ['blocked','BLOCKED BY ROBOTS.TXT', r => r.status === 'BLOCKED'],
    ['error','NO RESPONSE / ERROR', r => r.status === 'ERROR'],
    ['2xx','SUCCESS 2XX', r => { const n = parseInt(r.status); return n >= 200 && n < 300; }],
    ['3xx','REDIRECTION 3XX', null],   // rendered from redirects list
    ['chains','REDIRECT CHAINS', null],
    ['4xx','CLIENT ERROR 4XX', r => { const n = parseInt(r.status); return n >= 400 && n < 500; }],
    ['5xx','SERVER ERROR 5XX', r => parseInt(r.status) >= 500],
  ],
  url: [
    ['all','ALL', r => true],
    ['nonascii','NON-ASCII', r => r.url_flags.includes('non-ascii')],
    ['under','UNDERSCORES', r => r.url_flags.includes('underscores')],
    ['upper','UPPERCASE', r => r.url_flags.includes('uppercase')],
    ['params','PARAMETERS', r => r.url_flags.includes('parameters')],
    ['long','OVER 115 CHARS', r => r.url_flags.some(f => f.startsWith('over'))],
  ],
  title: [
    ['issues','ISSUES', r => r.issues.some(i => i.cat === 'title')],
    ['miss','MISSING', r => hasMsg(r,'title','Title missing')],
    ['dup','DUPLICATE', r => hasMsg(r,'title','Duplicate')],
    ['long','OVER 60 CHARS', r => hasMsg(r,'title','too long')],
    ['short','BELOW 30 CHARS', r => hasMsg(r,'title','too short')],
    ['sameh1','SAME AS H1', r => !!r.title && r.title === r.h1],
    ['all','ALL PAGES', r => r.is_html && r.status === '200'],
  ],
  desc: [
    ['issues','ISSUES', r => r.issues.some(i => i.cat === 'desc')],
    ['miss','MISSING', r => hasMsg(r,'desc','Description missing')],
    ['dup','DUPLICATE', r => hasMsg(r,'desc','Duplicate')],
    ['long','OVER 160 CHARS', r => hasMsg(r,'desc','too long')],
    ['short','BELOW 70 CHARS', r => hasMsg(r,'desc','too short')],
    ['all','ALL PAGES', r => r.is_html && r.status === '200'],
  ],
  keywords: [
    ['present','PRESENT', r => !!r.keywords],
    ['dup','DUPLICATE', r => r.kw_dup],
    ['miss','MISSING', r => r.is_html && r.status === '200' && !r.keywords],
    ['all','ALL PAGES', r => r.is_html && r.status === '200'],
  ],
  h1: [
    ['issues','ISSUES', r => r.issues.some(i => i.cat === 'h1')],
    ['miss','MISSING', r => hasMsg(r,'h1','H1 missing')],
    ['dup','DUPLICATE', r => hasMsg(r,'h1','Duplicate')],
    ['multi','MULTIPLE', r => hasMsg(r,'h1','Multiple')],
    ['long','OVER 70 CHARS', r => r.h1.length > 70],
    ['all','ALL PAGES', r => r.is_html && r.status === '200'],
  ],
  h2: [
    ['miss','MISSING', r => r.is_html && r.status === '200' && r.indexable && r.h2_count === 0],
    ['dup','DUPLICATE', r => r.h2_dup],
    ['multi','MULTIPLE', r => r.h2_count > 1],
    ['long','OVER 70 CHARS', r => r.h2.length > 70],
    ['all','ALL PAGES', r => r.is_html && r.status === '200'],
  ],
  content: [
    ['low','LOW CONTENT (<300 WORDS)', r => hasMsg(r,'thin','Thin content')],
    ['dupbody','EXACT DUPLICATES', r => r.dup_body],
    ['deep','DEEP PAGES (4+ CLICKS)', r => hasMsg(r,'depth','Deep')],
    ['noschema','NO STRUCTURED DATA', r => r.is_html && r.status === '200' && r.indexable && (!r.schema || !r.schema.length)],
    ['badschema','INVALID JSON-LD', r => hasMsg(r,'schema','Invalid')],
    ['nolang','MISSING LANG', r => hasMsg(r,'lang','lang')],
    ['all','ALL PAGES', r => r.is_html && r.status === '200'],
  ],
  canonical: [
    ['all','ALL PAGES', r => r.is_html && r.status === '200'],
    ['self','SELF-REFERENCING', r => !!r.canonical && r.canonical.replace(/\/$/,'') === r.url.replace(/\/$/,'')],
    ['canon','CANONICALISED', r => !r.indexable && r.index_reason.includes('Canonicalised')],
    ['miss','MISSING', r => hasMsg(r,'canonical','Canonical missing')],
    ['multi','MULTIPLE', r => r.canonical_count > 1],
  ],
  speed: [
    ['issues','ISSUES', r => r.issues.some(i => i.cat === 'speed')],
    ['slow','SLOW RESPONSE (>2S)', r => hasMsg(r,'speed','Slow')],
    ['large','LARGE HTML (>1MB)', r => hasMsg(r,'speed','Large')],
    ['many','MANY RESOURCES (>60)', r => hasMsg(r,'speed','Many')],
    ['all','ALL PAGES', r => r.is_html && r.status === '200'],
  ],
  internal: [
    ['all','ALL', x => true],
    ['html','HTML', x => x._kind === 'page' && x.is_html],
    ['img','IMAGES', x => x._kind === 'res' && x.type === 'image'],
    ['css','CSS', x => x._kind === 'res' && x.type === 'css'],
    ['js','JAVASCRIPT', x => x._kind === 'res' && x.type === 'js'],
    ['indexable','INDEXABLE', x => x._kind === 'page' && x.indexable],
    ['nonindex','NON-INDEXABLE', x => x._kind === 'page' && !x.indexable],
    ['broken','BROKEN', x => isBad(x.status)],
  ],
  external: [
    ['all','ALL', l => true],
    ['ok','2XX OK', l => { const n = parseInt(l.status); return n >= 200 && n < 300; }],
    ['3xx','REDIRECTING', l => { const n = parseInt(l.status); return n >= 300 && n < 400; }],
    ['broken','BROKEN / ERROR', l => isBad(l.status)],
  ],
  images: [
    ['all','ALL', r => true],
    ['noalt','MISSING ALT', r => r.missing_alt],
    ['large','LARGE (>200KB)', r => r.kb !== null && r.kb > 200],
    ['notwebp','NOT WEBP', r => ['jpg','png','gif','bmp'].includes(r.format)],
    ['nolazy','NO LAZY LOADING', r => r.lazy === 0],
    ['broken','BROKEN', r => isBad(r.status)],
  ],
};

function curSub(tab) {
  const defs = SUBDEFS[tab];
  if (!defs) return null;
  const sel = tabSub[tab] || defs[0][0];
  return defs.find(d => d[0] === sel) || defs[0];
}
function setTabSub(tab, id) { tabSub[tab] = id; render(); }

function internalRows() {
  const pageRows = snapshot.rows.map(r => ({...r, _kind:'page'}));
  const resRows = snapshot.resources.filter(x => x.internal).map(x => ({...x, _kind:'res'}));
  return pageRows.concat(resRows);
}
function datasetFor(tab) {
  if (tab === 'internal') return internalRows();
  if (tab === 'external') return snapshot.links.filter(l => !l.internal);
  if (tab === 'images') return snapshot.images;
  return snapshot.rows;
}
function filteredData(tab) {
  const sub = curSub(tab);
  const data = datasetFor(tab);
  if (!sub || !sub[2]) return data;
  return data.filter(sub[2]);
}
function subBar(tab, extraCounts) {
  const defs = SUBDEFS[tab];
  if (!defs) return '';
  const sel = tabSub[tab] || defs[0][0];
  const data = datasetFor(tab);
  return `<div class="subbar">` + defs.map(([id, label, fn]) => {
    const n = fn ? data.filter(fn).length : (extraCounts ? extraCounts[id] : 0);
    return `<div class="sub ${sel === id ? 'active' : ''}" onclick="setTabSub('${tab}','${id}')">${label}<span style="opacity:.65;margin-left:5px">${n}</span></div>`;
  }).join('') + `</div>`;
}

/* ---------------- crawl control ---------------- */
async function startCrawl() {
  const domain = document.getElementById('domain').value.trim();
  if (!domain) return;
  const maxPages = parseInt(document.getElementById('maxpages').value) || 500;
  let mode = document.getElementById('mode').value;
  if (domain.toLowerCase().endsWith('.xml')) {
    mode = 'sitemap';
    document.getElementById('mode').value = 'sitemap';
  }
  document.getElementById('startBtn').disabled = true;
  searchResults = null;
  await fetch('/crawl', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({url: domain, max_pages: maxPages, mode: mode})});
  document.getElementById('stopBtn').style.display = 'inline-block';
  document.getElementById('progress').style.display = 'flex';
  document.getElementById('tabs').style.display = 'flex';
  poll();
  pollTimer = setInterval(poll, 900);
}
async function stopCrawl() { await fetch('/stop', {method:'POST'}); }
async function poll() {
  const resp = await fetch('/status');
  snapshot = await resp.json();
  render();
  if (!snapshot.running && pollTimer) {
    clearInterval(pollTimer); pollTimer = null;
    document.getElementById('startBtn').disabled = false;
    document.getElementById('stopBtn').style.display = 'none';
    document.getElementById('progress').style.display = 'none';
  } else if (snapshot.running) {
    let txt;
    if (snapshot.phase === 'assets') {
      txt = `Phase 2/2: checking resources & external links — ${snapshot.assets_done}/${snapshot.assets_total}`;
    } else {
      txt = `Phase 1/2: crawled ${snapshot.counts.total} pages · ${snapshot.queued} in queue`;
    }
    document.getElementById('progressText').textContent = txt;
  }
}
function setFilter(f) {
  filter = f;
  document.querySelectorAll('.tab').forEach(t => t.classList.toggle('active', t.dataset.f === f));
  render();
}

function statusBadge(s) {
  if (s === '200') return `<span class="badge ok">200</span>`;
  if (s === 'ERROR') return `<span class="badge bad">ERR</span>`;
  if (s === 'BLOCKED') return `<span class="badge warn">🤖</span>`;
  if (s === '...' ) return `<span class="badge gray">…</span>`;
  if (s === 'not crawled') return `<span class="badge gray">n/c</span>`;
  const n = parseInt(s);
  if (n >= 400) return `<span class="badge bad">${s}</span>`;
  if (n >= 300) return `<span class="badge warn">${s}</span>`;
  return `<span class="badge ok">${esc(s)}</span>`;
}
function idxCell(r) {
  return r.indexable ?
    `<div style="font-size:.72rem;color:#0f6d33;margin-top:2px">indexable</div>` :
    `<div style="font-size:.72rem;color:#a11212;margin-top:2px">${esc(r.index_reason)}</div>`;
}
function issuesCell(r) {
  return r.issues.map(i => `<span class="issue ${i.cat}">${esc(i.msg)}</span>`).join('') ||
    '<span style="color:#0f6d33;font-weight:600">✓</span>';
}
function lenBadge(n, min, max) {
  const cls = n === 0 ? 'bad' : (n < min || n > max) ? 'warn' : 'ok';
  return `<span class="badge ${cls}" style="font-family:Consolas,monospace">${n}</span>`;
}

/* ---------------- SVG charts (no libraries) ---------------- */
function barSvg(buckets, title) {
  const max = Math.max(...buckets.map(b => b.value), 1);
  const n = buckets.length, W = n * 74 + 16, H = 152, chartH = 96, top = 24;
  const bars = buckets.map((b, i) => {
    const h = Math.max(2, Math.round(b.value / max * chartH));
    const x = 10 + i * 74, y = top + (chartH - h);
    return `<rect x="${x}" y="${y}" width="56" height="${h}" rx="4" fill="${b.color}"/>
      <text x="${x+28}" y="${y-5}" text-anchor="middle" font-size="12" font-weight="700" fill="#1a2b3c">${b.value}</text>
      <text x="${x+28}" y="${top+chartH+16}" text-anchor="middle" font-size="10" fill="#5a6b7c">${b.label}</text>`;
  }).join('');
  return `<div class="chartcard"><div class="charttitle">${title}</div>
    <svg viewBox="0 0 ${W} ${H}" style="width:100%;max-width:${Math.round(W*1.5)}px;height:auto;display:block">${bars}</svg></div>`;
}

function donutSvg(segments, title, centerLabel) {
  segments = segments.filter(s => s.value > 0);
  const total = segments.reduce((a, s) => a + s.value, 0) || 1;
  const r = 42, C = 2 * Math.PI * r;
  let offset = 0;
  const arcs = segments.map(s => {
    const len = s.value / total * C;
    const el = `<circle r="${r}" cx="60" cy="60" fill="none" stroke="${s.color}" stroke-width="20"
      stroke-dasharray="${len} ${C - len}" stroke-dashoffset="${-offset}" transform="rotate(-90 60 60)"/>`;
    offset += len;
    return el;
  }).join('');
  const legend = segments.map(s =>
    `<span><i style="background:${s.color}"></i>${s.label}: <b>${s.value}</b></span>`).join('');
  return `<div class="chartcard"><div class="charttitle">${title}</div>
    <div style="display:flex;align-items:center;gap:16px;flex-wrap:wrap">
      <svg viewBox="0 0 120 120" style="width:120px;height:120px;flex:0 0 auto">${arcs}
        <text x="60" y="58" text-anchor="middle" font-size="20" font-weight="800" fill="#1a2b3c">${total}</text>
        <text x="60" y="74" text-anchor="middle" font-size="9" fill="#5a6b7c">${centerLabel}</text>
      </svg>
      <div class="legend" style="flex-direction:column;align-items:flex-start;gap:5px">${legend}</div>
    </div></div>`;
}

function bucketize(values, edges, labels, colorFn) {
  const counts = new Array(labels.length).fill(0);
  values.forEach(v => {
    let i = edges.findIndex(e => v < e);
    if (i === -1) i = labels.length - 1;
    counts[i]++;
  });
  return labels.map((l, i) => ({label: l, value: counts[i], color: colorFn(i)}));
}

function depthChart() {
  const pages = snapshot.rows.filter(r => r.is_html && r.status === '200');
  if (snapshot.mode !== 'spider' || !pages.length) return '';
  const b = bucketize(pages.map(r => r.depth), [1,2,3,4], ['0 (home)','1','2','3','4+'],
    i => i >= 4 ? '#e0655a' : i === 3 ? '#e0b03d' : '#4d7fb0');
  return barSvg(b, 'Crawl Depth — clicks from home (healthy sites look like a pyramid)');
}
function inlinksChart() {
  const pages = snapshot.rows.filter(r => r.is_html && r.status === '200' && r.depth > 0);
  if (!pages.length) return '';
  const b = bucketize(pages.map(r => r.inlinks), [1,2,6,11,21], ['0','1','2–5','6–10','11–20','21+'],
    i => i <= 1 ? '#e0655a' : i === 2 ? '#e0b03d' : '#4caf7d');
  return barSvg(b, 'Internal Links per Page — 0–1 inlinks = near-orphan pages');
}
function wordsChart() {
  const pages = snapshot.rows.filter(r => r.is_html && r.status === '200' && r.words !== '');
  if (!pages.length) return '';
  const b = bucketize(pages.map(r => r.words), [100,300,600,1000,2000],
    ['<100','100–299','300–599','600–999','1000–1999','2000+'],
    i => i <= 1 ? '#e0655a' : i === 2 ? '#e0b03d' : '#4caf7d');
  return barSvg(b, 'Word Count Distribution — clustering under 300 = template-level thin content');
}
function titleLenChart() {
  const pages = snapshot.rows.filter(r => r.is_html && r.status === '200');
  if (!pages.length) return '';
  const b = bucketize(pages.map(r => r.title.length), [1,30,61,71],
    ['missing','1–29','30–60 ✓','61–70','71+'],
    i => i === 2 ? '#4caf7d' : i === 0 ? '#a11212' : '#e0b03d');
  return barSvg(b, 'Title Length Distribution — 30–60 chars displays fully in Google');
}
function responseDonut() {
  const rows = snapshot.rows;
  const seg = [
    {label: '2xx OK', color: '#4caf7d', value: rows.filter(r => { const n = parseInt(r.status); return n >= 200 && n < 300; }).length},
    {label: '3xx Redirect', color: '#e0b03d', value: snapshot.redirects.length},
    {label: '4xx Client Error', color: '#e0655a', value: rows.filter(r => { const n = parseInt(r.status); return n >= 400 && n < 500; }).length},
    {label: '5xx Server Error', color: '#a11212', value: rows.filter(r => parseInt(r.status) >= 500).length},
    {label: 'Blocked (robots)', color: '#b3541e', value: rows.filter(r => r.status === 'BLOCKED').length},
    {label: 'Error', color: '#8a99a8', value: rows.filter(r => r.status === 'ERROR').length},
  ];
  return donutSvg(seg, 'Response Codes', 'URLs');
}
function severityChart() {
  const d = snapshot.dashboard;
  if (!d) return '';
  const sum = s => d.works.filter(w => w.severity === s).reduce((a, w) => a + w.count, 0);
  const b = [
    {label: 'HIGH', value: sum('high'), color: '#e0655a'},
    {label: 'MEDIUM', value: sum('medium'), color: '#e0b03d'},
    {label: 'LOW', value: sum('low'), color: '#9db2c4'},
  ];
  return barSvg(b, 'Issues by Priority — total affected URLs per severity');
}

/* ---------------- render ---------------- */
function render() {
  if (!snapshot) return;
  const c = snapshot.counts;
  for (const k of ['internal','external','security','response','url','title','desc','keywords','h1','h2','content','images','canonical','speed','changes'])
    document.getElementById('n-'+k).textContent = c[k];
  const nbi = document.getElementById('nb-images');
  nbi.style.display = c.images_noalt ? 'inline-block' : 'none';
  nbi.textContent = c.images_noalt + ' no alt';
  const nbe = document.getElementById('nb-external');
  nbe.style.display = c.links_broken ? 'inline-block' : 'none';
  nbe.textContent = c.links_broken + ' ✗';

  const box = document.getElementById('tablebox');
  if (snapshot.error_note) { box.innerHTML = `<div class="empty">⚠ ${esc(snapshot.error_note)}</div>`; return; }

  if (filter === 'dashboard') { renderDashboard(box); return; }
  if (filter === 'internal') { renderInternal(box); return; }
  if (filter === 'external') { renderExternal(box); return; }
  if (filter === 'images') { renderImages(box); return; }
  if (filter === 'search') { renderSearch(box); return; }
  if (filter === 'linksugg') { renderLinkSugg(box); return; }
  if (filter === 'changes') { renderChanges(box); return; }
  if (filter === 'response') {
    const sel = tabSub['response'] || 'all';
    if (sel === '3xx' || sel === 'chains') { renderRedirects(box, sel); return; }
  }
  renderPageTable(box);
}

function renderPageTable(box) {
  const rows = filteredData(filter);
  const chainCount = snapshot.redirects.filter(r => r.hops > 1).length;
  let bar = subBar(filter, {'3xx': snapshot.redirects.length, 'chains': chainCount});
  if (filter === 'title') bar = `<div class="chartrow" style="grid-template-columns:1fr">${titleLenChart()}</div>` + bar;
  if (filter === 'content') bar = `<div class="chartrow" style="grid-template-columns:1fr">${wordsChart()}</div>` + bar;
  if (!rows.length) { box.innerHTML = bar + '<div class="empty">Nothing here 🎉</div>'; return; }

  let colHead, colFn;
  if (filter === 'title') {
    colHead = `<th style="width:34%">Meta Title</th><th style="width:6%">Chars</th>`;
    colFn = r => `<td>${esc(r.title) || '<span style="color:#a11212">(missing)</span>'}</td>
                  <td>${lenBadge(r.title.length, 30, 60)}</td>`;
  } else if (filter === 'desc') {
    colHead = `<th style="width:38%">Meta Description</th><th style="width:6%">Chars</th>`;
    colFn = r => `<td>${esc(r.description) || '<span style="color:#a11212">(missing)</span>'}</td>
                  <td>${lenBadge(r.description.length, 70, 160)}</td>`;
  } else if (filter === 'keywords') {
    colHead = `<th style="width:38%">Meta Keywords <span style="text-transform:none;font-weight:400">(ignored by Google — informational)</span></th><th style="width:6%">Chars</th>`;
    colFn = r => `<td>${esc(r.keywords) || '<span style="color:#8a99a8">(none)</span>'}
                  ${r.kw_dup ? '<div><span class="issue">Duplicate keywords</span></div>' : ''}</td>
                  <td>${r.keywords ? r.keywords.length : 0}</td>`;
  } else if (filter === 'h1') {
    colHead = `<th style="width:34%">H1 Content (all)</th><th style="width:8%">Count / Chars</th>`;
    colFn = r => `<td>${esc(r.h1_all) || '<span style="color:#a11212">(missing)</span>'}</td>
                  <td>${r.h1_count > 1 ? `<span class="badge warn">${r.h1_count} H1s</span>` :
                       r.h1_count === 0 ? `<span class="badge bad">0</span>` : `<span class="badge ok">1</span>`}
                      <div class="meta">${r.h1.length} chars</div></td>`;
  } else if (filter === 'h2') {
    colHead = `<th style="width:34%">H2 Content (first 5)</th><th style="width:8%">Count / Chars</th>`;
    colFn = r => `<td>${esc(r.h2_all) || '<span style="color:#a11212">(missing)</span>'}
                  ${r.h2_dup ? '<div><span class="issue h1">Duplicate first H2</span></div>' : ''}</td>
                  <td>${r.h2_count === 0 ? `<span class="badge bad">0</span>` :
                       `<span class="badge ok">${r.h2_count}</span>`}
                      <div class="meta">${r.h2.length} chars</div></td>`;
  } else if (filter === 'canonical') {
    colHead = `<th style="width:36%">Canonical URL</th><th style="width:8%">Match</th>`;
    colFn = r => `<td>${esc(r.canonical) || '<span style="color:#a11212">(missing)</span>'}
                  ${r.canonical_count > 1 ? `<div><span class="issue">${r.canonical_count} canonical tags!</span></div>` : ''}</td>
                  <td>${!r.canonical ? `<span class="badge bad">—</span>` :
                       r.canonical.replace(/\/$/,'') === r.url.replace(/\/$/,'') ?
                       `<span class="badge ok">self</span>` : `<span class="badge warn">other</span>`}</td>`;
  } else if (filter === 'url') {
    colHead = `<th style="width:8%">Length</th><th style="width:34%">URL Flags</th>`;
    colFn = r => `<td>${lenBadge(r.url.length, 1, 115)}</td>
                  <td>${r.url_flags.length ? r.url_flags.map(f => `<span class="issue">${esc(f)}</span>`).join('') :
                       '<span style="color:#0f6d33;font-weight:600">✓ clean</span>'}</td>`;
  } else if (filter === 'security') {
    colHead = `<th style="width:8%">Protocol</th><th style="width:34%">Security</th>`;
    colFn = r => `<td>${r.https ? '<span class="badge ok">https</span>' : '<span class="badge bad">http</span>'}</td>
                  <td>${(r.mixed > 0 ? `<span class="issue security">Mixed content (${r.mixed})</span>` : '') +
                       (r.sec_missing || []).map(h => `<span class="issue h1">no ${esc(h)}</span>`).join('') ||
                       '<span style="color:#0f6d33;font-weight:600">✓</span>'}</td>`;
  } else if (filter === 'speed') {
    colHead = `<th style="width:9%">Response</th><th style="width:9%">HTML Size</th><th style="width:9%">Resources</th>`;
    colFn = r => `<td>${r.ms === '' ? '—' : (r.ms > 2000 ? `<span class="badge warn">${r.ms}ms</span>` : `<span class="badge ok">${r.ms}ms</span>`)}</td>
                  <td>${r.kb === '' ? '—' : (r.kb > 1024 ? `<span class="badge warn">${r.kb}KB</span>` : `<span class="badge ok">${r.kb}KB</span>`)}</td>
                  <td>${r.res_count === '' ? '—' : (r.res_count > 60 ? `<span class="badge warn">${r.res_count}</span>` : `<span class="badge ok">${r.res_count}</span>`)}</td>`;
  } else if (filter === 'content') {
    colHead = `<th style="width:8%">Words</th><th style="width:10%">Depth</th><th style="width:20%">Content Flags</th>`;
    colFn = r => `<td>${r.words === '' ? '—' : (r.words < 300 ? `<span class="badge warn">${r.words}</span>` : `<span class="badge ok">${r.words}</span>`)}</td>
                  <td>${r.depth}</td>
                  <td>${(r.dup_body ? `<span class="issue thin">duplicate body (${r.dup_body_n} pages)</span>` : '') +
                       ((r.schema && r.schema.length) ? `<span class="type">${esc(r.schema.join(', '))}</span>` : '') +
                       (r.lang ? `<div class="meta">lang=${esc(r.lang)}</div>` : '<div class="meta" style="color:#a11212">no lang</div>')}</td>`;
  } else { // response and fallback
    colHead = `<th style="width:26%">Title</th><th style="width:12%">Linked From</th>`;
    colFn = r => `<td>${esc(r.title)}</td>
                  <td>${r.sources.length ? r.sources.map(s => `<div class="src">${esc(s)}</div>`).join('') : `<span class="meta">${r.inlinks} inlinks</span>`}</td>`;
  }

  box.innerHTML = bar + `<table><thead><tr><th style="width:30%">URL</th><th style="width:9%">Status / Index</th>
    ${colHead}<th>Issues</th></tr></thead><tbody>` +
    rows.map(r => `<tr data-url="${esc(r.url)}">
      <td class="urlcell"><a href="${esc(r.url)}" target="_blank">${esc(r.url)}</a></td>
      <td>${statusBadge(r.status)}${idxCell(r)}</td>
      ${colFn(r)}
      <td>${issuesCell(r)}</td></tr>`).join('') + '</tbody></table>';
}

function renderDashboard(box) {
  const d = snapshot.dashboard;
  const c = snapshot.counts;
  if (!d) { box.innerHTML = '<div class="empty">Dashboard appears once the crawl starts.</div>'; return; }
  const scoreColor = d.score >= 90 ? '#0f6d33' : d.score >= 75 ? '#8a6400' : d.score >= 60 ? '#b3541e' : '#a11212';
  const sevBadge = s => `<span class="badge ${s === 'high' ? 'bad' : s === 'medium' ? 'warn' : 'gray'}">${s.toUpperCase()}</span>`;
  const running = snapshot.running ? `<div class="meta" style="margin-top:6px">⏳ Crawl still running — score will update</div>` : '';

  box.innerHTML = `
  <div id="dashprint">
    <div class="dash-top">
      <div class="dash-score" style="border-color:${scoreColor}">
        <div class="dash-num" style="color:${scoreColor}">${d.score}</div>
        <div class="dash-grade">Grade ${d.grade}</div>
      </div>
      <div class="dash-meta">
        <div class="dash-site">${esc(snapshot.start_url)}</div>
        <div class="meta">${new Date().toISOString().slice(0,10)} · ${c.total} pages · ${c.images} images · ${c.external} external links${running}</div>
        <div class="dash-btns">
          <button onclick="location.href='/export/dashboard.xlsx'">⬇ Export Excel Report</button>
          <button class="pdfbtn" onclick="window.print()">🖨 Export PDF (Print)</button>
        </div>
      </div>
    </div>

    <div class="chartrow">${responseDonut()}${severityChart()}</div>

    <div class="dash-cols">
      <div class="dash-col">
        <div class="chghead">✅ What's Good (${d.goods.length})</div>
        ${d.goods.length ? d.goods.map(g => `<div class="dash-good">✓ ${esc(g)}</div>`).join('')
          : '<div class="src">Nothing passing yet</div>'}
      </div>
      <div class="dash-col">
        <div class="chghead">⚠️ Needs Work (${d.works.length})</div>
        ${d.works.length ? d.works.map((w, i) => `
          <div class="dash-work">
            <div class="dash-work-head" onclick="toggleWork(${i})">
              ${sevBadge(w.severity)}
              <b>${esc(w.label)}</b>
              <span class="badge gray">${w.count}</span>
              <span class="dash-expand" id="exp-${i}">▸</span>
            </div>
            <div class="dash-sugg">💡 ${esc(w.suggestion)}</div>
            <div class="dash-urls" id="urls-${i}" style="display:none">
              ${w.urls.slice(0, 50).map(u => `<div class="src"><a href="${esc(u.url)}" target="_blank" style="color:#1F4E79">${esc(u.url)}</a>${u.detail ? ` <span style="color:#8a99a8">— ${esc(u.detail)}</span>` : ''}</div>`).join('')}
              ${w.count > 50 ? `<div class="src">+ ${w.count - 50} more — see the relevant tab or the Excel report</div>` : ''}
            </div>
          </div>`).join('')
          : '<div class="dash-good" style="font-weight:700">🎉 Nothing needs work — clean audit!</div>'}
      </div>
    </div>
    <div style="text-align:center;margin-top:18px;font-size:.74rem;color:#8a99a8">
      Generated by <b>Makdi</b> 🕷 — the free unlimited SEO crawler ·
      <a href="https://vikasdisale.com/makdi" target="_blank" style="color:#1F4E79">vikasdisale.com/makdi</a>
    </div>
  </div>`;
}
function toggleWork(i) {
  const el = document.getElementById('urls-' + i);
  const ex = document.getElementById('exp-' + i);
  const open = el.style.display !== 'none';
  el.style.display = open ? 'none' : 'block';
  ex.textContent = open ? '▸' : '▾';
}

function renderInternal(box) {
  const rows = filteredData('internal');
  const charts = (depthChart() || inlinksChart()) ?
    `<div class="chartrow">${depthChart()}${inlinksChart()}</div>` : '';
  const bar = charts + subBar('internal');
  if (!rows.length) { box.innerHTML = bar + '<div class="empty">Nothing here 🎉</div>'; return; }
  box.innerHTML = bar + `<table><thead><tr><th style="width:38%">URL</th><th style="width:8%">Type</th>
    <th style="width:9%">Status</th><th style="width:16%">Indexability</th><th style="width:7%">Inlinks</th><th>Issues / Found On</th></tr></thead><tbody>` +
    rows.map(x => {
      if (x._kind === 'page') {
        return `<tr data-url="${esc(x.url)}"><td class="urlcell"><a href="${esc(x.url)}" target="_blank">${esc(x.url)}</a></td>
          <td><span class="type">${x.is_html ? 'html' : 'other'}</span></td>
          <td>${statusBadge(x.status)}</td>
          <td style="font-size:.8rem;color:${x.indexable ? '#0f6d33' : '#a11212'}">${x.indexable ? 'Indexable' : esc(x.index_reason)}</td>
          <td><b>${x.inlinks}</b></td>
          <td>${issuesCell(x)}</td></tr>`;
      }
      return `<tr data-url="${esc(x.url)}"><td class="urlcell"><a href="${esc(x.url)}" target="_blank">${esc(x.url)}</a></td>
        <td><span class="type">${x.type}</span></td>
        <td>${statusBadge(x.status)}</td>
        <td style="font-size:.8rem;color:#8a99a8">resource</td>
        <td><b>${x.inlinks}</b></td>
        <td>${x.sources.map(s => `<div class="src">${esc(s)}</div>`).join('')}</td></tr>`;
    }).join('') + '</tbody></table>';
}

function renderExternal(box) {
  const rows = [...filteredData('external')].sort((a,b) => b.count - a.count);
  const bar = subBar('external');
  if (!rows.length) { box.innerHTML = bar + '<div class="empty">Nothing here 🎉</div>'; return; }
  box.innerHTML = bar + `<table><thead><tr><th style="width:36%">External URL</th>
    <th style="width:8%">Status</th><th style="width:8%">Inlinks</th><th>Found On (with anchor text)</th></tr></thead><tbody>` +
    rows.map(l =>
      `<tr data-url="${esc(l.target)}"><td class="urlcell"><a href="${esc(l.target)}" target="_blank">${esc(l.target)}</a></td>
       <td>${statusBadge(l.status)}</td>
       <td><b>${l.count}</b></td>
       <td>${l.examples.map(e => `<div class="src">${esc(e.source)} — <span class="anchor">"${esc(e.anchor)}"</span></div>`).join('')}
       ${l.count > l.examples.length ? `<div class="src">+ ${l.count - l.examples.length} more</div>` : ''}</td></tr>`
    ).join('') + '</tbody></table>';
}

function renderImages(box) {
  const rows = filteredData('images');
  const bar = subBar('images');
  if (!rows.length) { box.innerHTML = bar + '<div class="empty">Nothing here 🎉</div>'; return; }
  box.innerHTML = bar + `<table><thead><tr><th style="width:32%">Image URL</th>
    <th style="width:7%">Format</th><th style="width:8%">Size</th><th style="width:7%">Status</th>
    <th style="width:24%">Alt Text</th><th>Found On</th></tr></thead><tbody>` +
    rows.map(r => {
      const kb = r.kb === null ? '<span class="badge gray">?</span>' :
        r.kb > 200 ? `<span class="badge warn">${r.kb} KB</span>` : `<span class="badge ok">${r.kb} KB</span>`;
      const fmt = ['webp','avif','svg'].includes(r.format) ? `<span class="badge ok">${r.format}</span>` :
        r.format === '?' ? `<span class="badge gray">?</span>` : `<span class="badge warn">${r.format}</span>`;
      const alt = r.missing_alt ?
        `<span class="badge bad">missing</span>` +
        (r.alts.length ? `<div class="src">also used with: ${r.alts.map(a=>`"${esc(a)}"`).join(', ')}</div>` : '') :
        r.alts.map(a => `<div class="anchor">"${esc(a)}"</div>`).join('');
      return `<tr data-url="${esc(r.url)}">
        <td class="urlcell"><a href="${esc(r.url)}" target="_blank">${esc(r.url)}</a><div class="meta">used ${r.uses}×</div></td>
        <td>${fmt}<div class="meta">${r.lazy > 0 ? 'lazy ✓' : 'no lazy'}</div></td>
        <td>${kb}</td>
        <td>${statusBadge(r.status)}</td>
        <td>${alt}</td>
        <td>${r.sources.map(s => `<div class="src">${esc(s)}</div>`).join('')}</td></tr>`;
    }).join('') + '</tbody></table>';
}

function renderRedirects(box, sel) {
  const chainCount = snapshot.redirects.filter(r => r.hops > 1).length;
  const bar = subBar('response', {'3xx': snapshot.redirects.length, 'chains': chainCount});
  const rows = sel === 'chains' ? snapshot.redirects.filter(r => r.hops > 1) : snapshot.redirects;
  if (!rows.length) { box.innerHTML = bar + '<div class="empty">No redirects here 🎉</div>'; return; }
  box.innerHTML = bar + `<table><thead><tr><th style="width:26%">Original URL (redirecting)</th><th style="width:6%">Type</th>
    <th style="width:24%">Redirects To (final)</th><th style="width:24%">Found On (fix link here)</th><th>Chain</th></tr></thead><tbody>` +
    rows.map(r =>
      `<tr data-url="${esc(r.from)}"><td class="urlcell"><a href="${esc(r.from)}" target="_blank">${esc(r.from)}</a></td>
       <td>${statusBadge(String(r.first_status))}${r.hops > 1 ? ' <span class="badge warn">'+r.hops+' hops</span>' : ''}</td>
       <td class="urlcell"><a href="${esc(r.to)}" target="_blank">${esc(r.to)}</a></td>
       <td>${r.sources && r.sources.length ?
            r.sources.map(s => `<div class="src">${esc(s)}</div>`).join('') :
            '<span style="color:#8a99a8;font-size:.78rem">(direct / sitemap entry)</span>'}</td>
       <td class="chain">${r.chain.map(x => esc(x.url)+' <b>['+x.status+']</b>').join(' → ')}</td></tr>`
    ).join('') + '</tbody></table>';
}

function renderSearch(box) {
  const bar = `<div class="searchbar">
    <select id="searchMode" onchange="searchMode=this.value">
      <option value="contains" ${searchMode==='contains'?'selected':''}>CONTAINS</option>
      <option value="not" ${searchMode==='not'?'selected':''}>DOES NOT CONTAIN</option>
    </select>
    <input type="text" id="searchQ" placeholder="Search page source, e.g. gtag( or UA-12345 or any text" value="${esc(searchQ)}"
      onkeydown="if(event.key==='Enter')runSearch()">
    <button onclick="runSearch()">Search</button>
  </div>`;
  let body;
  if (searchResults === null) {
    body = '<div class="empty">Search the raw HTML source of every crawled page.<br>Examples: find pages missing your analytics tag (DOES NOT CONTAIN "gtag"), pages mentioning an old brand name, or pages containing a shortcode.</div>';
  } else if (!searchResults.length) {
    body = '<div class="empty">No matching pages.</div>';
  } else if (searchMode === 'not') {
    body = `<table><thead><tr><th>Pages NOT containing "${esc(searchQ)}" (${searchResults.length})</th></tr></thead><tbody>` +
      searchResults.map(r => `<tr><td class="urlcell"><a href="${esc(r.url)}" target="_blank">${esc(r.url)}</a></td></tr>`).join('') +
      '</tbody></table>';
  } else {
    body = `<table><thead><tr><th style="width:36%">URL</th><th style="width:9%">Matches</th><th>Snippet</th></tr></thead><tbody>` +
      searchResults.map(r =>
        `<tr><td class="urlcell"><a href="${esc(r.url)}" target="_blank">${esc(r.url)}</a></td>
         <td><b>${r.count}</b></td>
         <td style="font-size:.8rem;color:#5a6b7c">…${esc(r.snippet)}…</td></tr>`).join('') +
      '</tbody></table>';
  }
  box.innerHTML = bar + body;
}

async function runSearch() {
  const q = document.getElementById('searchQ').value.trim();
  if (q.length < 2) return;
  searchQ = q;
  searchMode = document.getElementById('searchMode').value;
  const resp = await fetch('/search', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({q: q, mode: searchMode})});
  const data = await resp.json();
  searchResults = data.results;
  render();
}

function renderChanges(box) {
  const ch = snapshot.changes;
  if (!ch) { box.innerHTML = '<div class="empty">Changes appear after the crawl finishes.</div>'; return; }
  if (ch.first_crawl) {
    box.innerHTML = '<div class="empty">✅ First crawl of this domain saved.<br>Run another crawl later and this tab will show exactly what changed.</div>'; return;
  }
  let h = `<div class="empty" style="text-align:left;padding:16px 20px">Compared with previous crawl: <b>${esc(ch.prev_date)}</b></div>`;
  const section = (title, items, fmt) => {
    h += `<div class="chghead">${title} (${items.length})</div>`;
    if (!items.length) { h += `<div class="src" style="margin-left:4px">none</div>`; return; }
    h += `<table><tbody>` + items.map(fmt).join('') + `</tbody></table>`;
  };
  section(`🔴 New issues since last crawl`, ch.new_issues,
    i => `<tr><td class="urlcell" style="width:50%"><a href="${esc(i.url)}" target="_blank">${esc(i.url)}</a></td><td><span class="chgnew">${esc(i.msg)}</span></td></tr>`);
  section(`🟢 Fixed / resolved issues`, ch.resolved,
    i => `<tr><td class="urlcell" style="width:50%"><a href="${esc(i.url)}" target="_blank">${esc(i.url)}</a></td><td><span class="chgres">${esc(i.msg)}</span></td></tr>`);
  section(`🆕 New pages`, ch.new_pages,
    u => `<tr><td class="urlcell"><a href="${esc(u)}" target="_blank">${esc(u)}</a></td></tr>`);
  section(`🗑 Removed pages (or beyond crawl limit)`, ch.removed_pages,
    u => `<tr><td class="urlcell">${esc(u)}</td></tr>`);
  box.innerHTML = h;
}

/* ---------------- URL detail panel (Screaming Frog style) ---------------- */
document.addEventListener('click', e => {
  const tr = e.target.closest && e.target.closest('tr[data-url]');
  if (tr && !e.target.closest('a') && !e.target.closest('button')) {
    showDetails(tr.dataset.url);
  }
});
document.addEventListener('keydown', e => { if (e.key === 'Escape') closeDetails(); });

function closeDetails() {
  document.getElementById('overlay').style.display = 'none';
  document.getElementById('detail').style.display = 'none';
}

async function showDetails(url) {
  const panel = document.getElementById('detail');
  document.getElementById('overlay').style.display = 'block';
  panel.style.display = 'block';
  panel.innerHTML = `<button class="detail-close" onclick="closeDetails()">✕ close</button>
    <h2>${esc(url)}</h2><div class="empty">Loading…</div>`;
  let d;
  try {
    const resp = await fetch('/urldata?url=' + encodeURIComponent(url));
    d = await resp.json();
  } catch (e) { d = {inlinks: [], outlinks: [], images: [], resources: []}; }
  renderDetailPanel(url, d);
}

function renderDetailPanel(url, d) {
  const panel = document.getElementById('detail');
  const r = snapshot.rows.find(x => x.url === url);
  const img = snapshot.images.find(x => x.url === url);
  const res = snapshot.resources.find(x => x.url === url);
  const ext = snapshot.links.find(x => x.target === url && !x.internal);
  const redir = snapshot.redirects.find(x => x.from === url);

  let head = `<button class="detail-close" onclick="closeDetails()">✕ close</button>
    <h2>${esc(url)}</h2>
    <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">`;
  if (r) head += statusBadge(r.status) +
    (r.indexable ? '<span class="badge ok">Indexable</span>'
                 : `<span class="badge bad">${esc(r.index_reason)}</span>`);
  else if (img) head += statusBadge(img.status) + '<span class="type">image</span>';
  else if (res) head += statusBadge(res.status) + `<span class="type">${res.type}</span>`;
  else if (ext) head += statusBadge(ext.status) + '<span class="type">external</span>';
  head += ` <a href="${esc(url)}" target="_blank" style="color:#1F4E79;font-size:.85rem;font-weight:700">open ↗</a></div>`;

  let body = '';

  if (r) {
    const ov = [
      ['Title', r.title ? `${esc(r.title)} <span class="meta">(${r.title.length} chars)</span>` : '<span style="color:#a11212">missing</span>'],
      ['Description', r.description ? `${esc(r.description)} <span class="meta">(${r.description.length} chars)</span>` : '<span style="color:#a11212">missing</span>'],
      ['Keywords', r.keywords ? esc(r.keywords) : '<span class="meta">none</span>'],
      ['H1', r.h1_all ? `${esc(r.h1_all)} <span class="meta">(${r.h1_count})</span>` : '<span style="color:#a11212">missing</span>'],
      ['H2', r.h2_all ? `${esc(r.h2_all)} <span class="meta">(${r.h2_count})</span>` : '<span class="meta">none</span>'],
      ['Canonical', r.canonical ? esc(r.canonical) + (r.canonical_count > 1 ? ` <span style="color:#a11212">(${r.canonical_count} tags!)</span>` : '') : '<span style="color:#a11212">missing</span>'],
      ['Word count', r.words === '' ? '—' : r.words],
      ['Crawl depth', r.depth + ' clicks from home'],
      ['Inlinks', r.inlinks],
      ['Response', r.ms === '' ? '—' : r.ms + ' ms · ' + r.kb + ' KB · ' + r.res_count + ' resources'],
      ['Language', r.lang ? esc(r.lang) : '<span style="color:#a11212">no lang attribute</span>'],
      ['Schema', r.schema && r.schema.length ? esc(r.schema.join(', ')) : '<span class="meta">none</span>'],
      ['Security', (r.https ? 'https' : '<b style="color:#a11212">http</b>') +
        (r.mixed ? ` · <span style="color:#a11212">${r.mixed} mixed content</span>` : '') +
        (r.sec_missing && r.sec_missing.length ? ` · missing: ${esc(r.sec_missing.join(', '))}` : '')],
      ['URL flags', r.url_flags.length ? esc(r.url_flags.join(', ')) : '<span class="meta">clean</span>'],
    ];
    body += `<div class="detail-sec">Overview</div><table class="ov">` +
      ov.map(([k, v]) => `<tr><td>${k}</td><td>${v}</td></tr>`).join('') + `</table>`;
    if (r.issues.length)
      body += `<div class="detail-sec">Issues (${r.issues.length})</div>` +
        r.issues.map(i => `<span class="issue ${i.cat}">${esc(i.msg)}</span>`).join(' ');
  }
  if (img) {
    body += `<div class="detail-sec">Image Data</div><table class="ov">
      <tr><td>Format / Size</td><td>${esc(img.format)} · ${img.kb === null ? '?' : img.kb + ' KB'}</td></tr>
      <tr><td>Alt text</td><td>${img.missing_alt ? '<span style="color:#a11212">missing on some pages</span> ' : ''}${img.alts.map(a => `"${esc(a)}"`).join(', ')}</td></tr>
      <tr><td>Lazy loading</td><td>${img.lazy > 0 ? 'yes' : 'no'}</td></tr>
      <tr><td>Used</td><td>${img.uses}× across the site</td></tr></table>`;
  }
  if (redir) {
    body += `<div class="detail-sec">Redirect</div>
      <div class="dl-item">${statusBadge(String(redir.first_status))} → ${esc(redir.to)} (${redir.hops} hop${redir.hops > 1 ? 's' : ''})</div>
      <div class="chain" style="margin-top:6px">${redir.chain.map(x => esc(x.url) + ' <b>[' + x.status + ']</b>').join(' → ')}</div>`;
  }

  body += `<div class="detail-sec">Inlinks — pages linking here (${d.inlinks.length})</div>`;
  body += d.inlinks.length ?
    d.inlinks.map(l => `<div class="dl-item"><a href="#" onclick="showDetails('${esc(l.source)}');return false">${esc(l.source)}</a>
      <div class="meta">anchor: <span class="anchor">"${esc(l.anchor)}"</span></div></div>`).join('') :
    `<div class="meta">No internal links point here${r && r.depth === 0 ? ' (start URL)' : ' — orphan risk'}</div>`;

  if (d.outlinks.length) {
    body += `<div class="detail-sec">Outlinks — links from this page (${d.outlinks.length})</div>` +
      d.outlinks.map(l => `<div class="dl-item">${statusBadge(l.status)}
        <span class="type">${l.internal ? 'int' : 'ext'}</span>
        <a href="#" onclick="showDetails('${esc(l.target)}');return false">${esc(l.target)}</a>
        <div class="meta">anchor: <span class="anchor">"${esc(l.anchor)}"</span></div></div>`).join('');
  }

  if (d.images.length) {
    body += `<div class="detail-sec">Images on this page (${d.images.length})</div>` +
      d.images.map(i => `<div class="dl-item">${statusBadge(i.status)}
        ${i.kb !== null ? `<span class="badge ${i.kb > 200 ? 'warn' : 'ok'}">${i.kb} KB</span>` : ''}
        ${i.missing_alt ? '<span class="badge bad">no alt</span>' : ''}
        ${esc(i.url)}</div>`).join('');
  }

  if (d.resources.length) {
    body += `<div class="detail-sec">CSS / JS on this page (${d.resources.length})</div>` +
      d.resources.map(x => `<div class="dl-item">${statusBadge(x.status)} <span class="type">${x.type}</span> ${esc(x.url)}</div>`).join('');
  }

  panel.innerHTML = head + body;
  panel.scrollTop = 0;
}

function renderLinkSugg(box) {
  const bar = `<div class="subbar">
    <div class="sub ${linkSuggFilter==='all'?'active':''}" onclick="linkSuggFilter='all';render()">ALL SUGGESTIONS</div>
    <div class="sub ${linkSuggFilter==='mention'?'active':''}" onclick="linkSuggFilter='mention';render()">EXACT MENTIONS ⭐</div>
    <div class="sub ${linkSuggFilter==='topic'?'active':''}" onclick="linkSuggFilter='topic';render()">TOPIC MATCHES</div>
    <div class="sub" style="margin-left:auto;background:#1F4E79;border-color:#1F4E79;color:#fff" onclick="genLinkSugg()">${linkSugg?'↻ Regenerate':'⚡ Generate suggestions'}</div>
  </div>`;
  if (linkSuggLoading) { box.innerHTML = bar + '<div class="empty"><div class="spinner" style="margin:0 auto 12px"></div>Analyzing page content and link graph…</div>'; return; }
  if (!linkSugg) {
    box.innerHTML = bar + `<div class="empty">Finds internal linking opportunities two ways:<br><br>
      <b>⭐ Exact mentions</b> — a page's text mentions another page's topic but doesn't link to it (the anchor text is already written!)<br>
      <b>Topic matches</b> — pages sharing 5+ significant keywords with no link between them<br><br>
      Prioritised toward your weakest pages (fewest inlinks). Click <b>⚡ Generate suggestions</b> after a crawl.</div>`;
    return;
  }
  const groups = linkSugg.targets.map(t => ({...t,
      suggestions: t.suggestions.filter(s => linkSuggFilter==='all' || s.type===linkSuggFilter)}))
    .filter(t => t.suggestions.length);
  if (!groups.length) { box.innerHTML = bar + '<div class="empty">No suggestions in this filter — your internal linking looks healthy 🎉</div>'; return; }
  box.innerHTML = bar +
    `<div class="empty" style="text-align:left;padding:12px 18px">Analyzed <b>${linkSugg.pages_analyzed}</b> pages · <b>${linkSugg.mentions}</b> exact unlinked mentions found. Weakest pages first.</div>` +
    groups.map(t => `
    <div class="dash-work" style="background:#fff">
      <div class="dash-work-head" style="cursor:default">
        <span class="badge ${t.inlinks<=1?'bad':'warn'}">${t.inlinks} inlinks</span>
        <b>Link TO:</b> <a href="${esc(t.target)}" target="_blank" style="color:#1F4E79">${esc(t.target)}</a>
      </div>
      <div class="meta" style="margin:2px 0 8px">${esc(t.title)}</div>
      ${t.suggestions.map(s => `
        <div class="dl-item">
          ${s.type==='mention' ? '<span class="badge ok">⭐ mention</span>' : '<span class="badge gray">topic</span>'}
          <b>from</b> <a href="${esc(s.source)}" target="_blank" style="color:#1F4E79">${esc(s.source)}</a>
          <div class="meta">suggested anchor: <span class="anchor">"${esc(s.anchor)}"</span> · ${esc(s.evidence)}</div>
        </div>`).join('')}
    </div>`).join('');
}
async function genLinkSugg() {
  linkSuggLoading = true; render();
  try {
    const resp = await fetch('/linksuggest');
    linkSugg = await resp.json();
  } catch (e) { linkSugg = {targets:[], mentions:0, pages_analyzed:0}; }
  linkSuggLoading = false; render();
}

/* ---------------- copy & export ---------------- */
function currentUrls() {
  if (filter === 'external') return filteredData('external').map(l => l.target);
  if (filter === 'images') return filteredData('images').map(r => r.url);
  if (filter === 'search') return (searchResults || []).map(r => r.url);
  if (filter === 'linksugg') return linkSugg ? [...new Set(linkSugg.targets.map(t => t.target))] : [];
  if (filter === 'linksugg') {
    if (!linkSugg) return;
    const rows = [];
    linkSugg.targets.forEach(t => t.suggestions.forEach(s => {
      if (linkSuggFilter==='all' || s.type===linkSuggFilter)
        rows.push([t.target, t.inlinks, s.source, s.type, s.anchor, s.evidence]);
    }));
    downloadCsv(['Link To (target)','Target Inlinks','Add Link From (source)','Type','Suggested Anchor','Evidence'], rows, 'link_suggestions');
    return;
  }
  if (filter === 'changes') {
    const ch = snapshot.changes;
    return ch && ch.new_issues ? [...new Set(ch.new_issues.map(i => i.url))] : [];
  }
  if (filter === 'response') {
    const sel = tabSub['response'] || 'all';
    if (sel === '3xx') return snapshot.redirects.map(r => r.from);
    if (sel === 'chains') return snapshot.redirects.filter(r => r.hops > 1).map(r => r.from);
  }
  return filteredData(filter).map(x => x.url);
}
function copyUrls() {
  if (!snapshot) return;
  const urls = currentUrls();
  if (!urls.length) return;
  navigator.clipboard.writeText(urls.join('\n')).then(() => {
    const btns = document.querySelectorAll('.tab');
    const btn = btns[btns.length - 2];
    const old = btn.textContent;
    btn.textContent = `✓ Copied ${urls.length} URLs`;
    setTimeout(() => { btn.textContent = old; }, 1500);
  });
}

function csvCell(v) {
  v = String(v ?? '');
  return (v.includes(',') || v.includes('"') || v.includes('\n'))
    ? '"' + v.replace(/"/g, '""') + '"' : v;
}
function downloadCsv(headers, rows, name) {
  const lines = [headers.map(csvCell).join(',')]
    .concat(rows.map(r => r.map(csvCell).join(',')));
  const blob = new Blob(['\ufeff' + lines.join('\r\n')], {type: 'text/csv;charset=utf-8'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  const domain = (snapshot.start_url || 'site').replace(/https?:\/\//,'').replace(/[^a-z0-9.]/gi,'_').replace(/\.+$/,'');
  a.download = `${domain}_${name}_${new Date().toISOString().slice(0,10)}.csv`;
  a.click();
  URL.revokeObjectURL(a.href);
}
function exportCsv() {
  if (!snapshot) return;
  const sel = tabSub[filter] || (SUBDEFS[filter] ? SUBDEFS[filter][0][0] : 'all');
  const name = filter + (sel !== 'all' && sel !== 'issues' ? '_' + sel : '');

  if (filter === 'response' && (sel === '3xx' || sel === 'chains')) {
    const rows = sel === 'chains' ? snapshot.redirects.filter(r => r.hops > 1) : snapshot.redirects;
    downloadCsv(['From URL','Type','Hops','Final URL','Found On','Full Chain'],
      rows.map(r => [r.from, r.first_status, r.hops, r.to,
        (r.sources || []).join(' | '),
        r.chain.map(x => `${x.url} [${x.status}]`).join(' -> ')]), name);
    return;
  }
  if (filter === 'internal') {
    downloadCsv(['URL','Kind','Type','Status','Indexability','Inlinks','Issues'],
      filteredData('internal').map(x => x._kind === 'page' ?
        [x.url, 'page', x.is_html ? 'html' : 'other', x.status,
         x.indexable ? 'Indexable' : x.index_reason, x.inlinks,
         x.issues.map(i => i.msg).join(' | ')] :
        [x.url, 'resource', x.type, x.status, '', x.inlinks, '']), name);
    return;
  }
  if (filter === 'external') {
    downloadCsv(['External URL','Status','Inlinks','Example Sources (with anchor)'],
      filteredData('external').map(l => [l.target, l.status, l.count,
        l.examples.map(e => `${e.source} ["${e.anchor}"]`).join(' | ')]), name);
    return;
  }
  if (filter === 'images') {
    downloadCsv(['Image URL','Format','Size KB','Status','Alt Missing','Alt Texts','Lazy Loading','Times Used','Found On'],
      filteredData('images').map(r => [r.url, r.format, r.kb === null ? '' : r.kb, r.status,
        r.missing_alt ? 'YES' : 'no', r.alts.join(' | '), r.lazy > 0 ? 'yes' : 'no', r.uses,
        r.sources.join(' | ')]), name);
    return;
  }
  if (filter === 'search') {
    downloadCsv(['URL','Matches','Snippet'],
      (searchResults || []).map(r => [r.url, r.count, r.snippet]),
      'search_' + searchMode);
    return;
  }
  if (filter === 'linksugg') {
    if (!linkSugg) return;
    const rows = [];
    linkSugg.targets.forEach(t => t.suggestions.forEach(s => {
      if (linkSuggFilter==='all' || s.type===linkSuggFilter)
        rows.push([t.target, t.inlinks, s.source, s.type, s.anchor, s.evidence]);
    }));
    downloadCsv(['Link To (target)','Target Inlinks','Add Link From (source)','Type','Suggested Anchor','Evidence'], rows, 'link_suggestions');
    return;
  }
  if (filter === 'changes') {
    const ch = snapshot.changes;
    if (!ch || ch.first_crawl) return;
    const rows = []
      .concat(ch.new_issues.map(i => ['NEW ISSUE', i.url, i.msg]))
      .concat(ch.resolved.map(i => ['RESOLVED', i.url, i.msg]))
      .concat(ch.new_pages.map(u => ['NEW PAGE', u, '']))
      .concat(ch.removed_pages.map(u => ['REMOVED PAGE', u, '']));
    downloadCsv(['Change Type','URL','Issue'], rows, 'changes');
    return;
  }
  // generic page export
  downloadCsv(
    ['URL','Status','Indexability','Depth','Inlinks','Words','Response ms','Size KB','Resources',
     'Protocol','Mixed Content','Missing Security Headers','URL Flags',
     'Lang','Schema Types','Title','Title Length','Meta Description','Description Length',
     'Meta Keywords','H1 (all)','H1 Count','H2 (first 5)','H2 Count','Canonical','Canonical Count','Issues'],
    filteredData(filter).map(r => [
      r.url, r.status, r.indexable ? 'Indexable' : r.index_reason,
      r.depth, r.inlinks, r.words, r.ms, r.kb, r.res_count,
      r.https ? 'https' : 'http', r.mixed, (r.sec_missing || []).join(' | '),
      r.url_flags.join(' | '),
      r.lang, (r.schema || []).join(' | '),
      r.title, r.title ? r.title.length : 0,
      r.description, r.description ? r.description.length : 0,
      r.keywords, r.h1_all, r.h1_count, r.h2_all, r.h2_count,
      r.canonical, r.canonical_count,
      r.issues.map(i => i.msg).join(' | ')]), name);
}
</script>
</body>
</html>"""


class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a):
        pass

    def _json(self, obj, code=200):
        body = json.dumps(obj).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        if self.path == "/status":
            self._json(build_snapshot())
            return
        if self.path.startswith("/urldata?"):
            from urllib.parse import parse_qs
            q = parse_qs(self.path.split("?", 1)[1])
            url = (q.get("url") or [""])[0]
            self._json(url_detail(url))
            return
        if self.path == "/linksuggest":
            self._json(link_suggestions())
            return
        if self.path == "/export/dashboard.xlsx":
            self._export_xlsx()
            return
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.end_headers()
        self.wfile.write(HTML.encode("utf-8"))

    def _export_xlsx(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            self._json({"error": "openpyxl not installed. Run: py -m pip install openpyxl"}, 500)
            return
        import io
        snap = build_snapshot()
        dash = snap["dashboard"]
        c = snap["counts"]

        wb = Workbook()
        navy = PatternFill("solid", start_color="1F4E79")
        white_bold = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        body = Font(name="Arial", size=10)
        h_font = Font(name="Arial", size=12, bold=True, color="1F4E79")
        sev_fill = {"high": PatternFill("solid", start_color="FFC7CE"),
                    "medium": PatternFill("solid", start_color="FFEB9C"),
                    "low": PatternFill("solid", start_color="DDEBF7")}

        # ---- Sheet 1: Dashboard ----
        ws = wb.active
        ws.title = "Dashboard"
        ws.column_dimensions["A"].width = 58
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 90
        rows_out = [
            ("SEO AUDIT DASHBOARD", "", "", ""),
            (f"Site: {snap['start_url']}", "", "", ""),
            (f"Date: {time.strftime('%Y-%m-%d %H:%M')}", "", "", ""),
            (f"Health Score: {dash['score']} / 100  (Grade {dash['grade']})", "", "", ""),
            (f"Pages crawled: {c['total']}   Images: {c['images']}   External links: {c['external']}", "", "", ""),
            ("", "", "", ""),
            ("WHAT'S GOOD", "", "", ""),
        ]
        for g in dash["goods"]:
            rows_out.append(("  ✓ " + g, "", "", ""))
        rows_out.append(("", "", "", ""))
        rows_out.append(("NEEDS WORK", "Count", "Priority", "Suggestion"))
        header_row_idx = len(rows_out)
        for w in dash["works"]:
            rows_out.append(("  " + w["label"], w["count"], w["severity"].upper(), w["suggestion"]))
        for row in rows_out:
            ws.append(row)
        ws.append(("", "", "", ""))
        ws.append((f"Generated by {TOOL_NAME} v{VERSION} — the free unlimited SEO crawler · {TOOL_URL}", "", "", ""))
        ws.cell(row=ws.max_row, column=1).font = Font(name="Arial", size=9, italic=True, color="8A99A8")
        for r in range(1, ws.max_row + 1):
            for col in range(1, 5):
                ws.cell(row=r, column=col).font = body
                ws.cell(row=r, column=col).alignment = Alignment(vertical="top", wrap_text=(col == 4))
        for r, txt in [(1, None), (7, None), (header_row_idx, None)]:
            ws.cell(row=r, column=1).font = h_font
        for col in range(1, 5):
            cell = ws.cell(row=header_row_idx, column=col)
            cell.fill = navy
            cell.font = white_bold
        for i, w in enumerate(dash["works"]):
            ws.cell(row=header_row_idx + 1 + i, column=3).fill = sev_fill[w["severity"]]
        ws.freeze_panes = "A2"

        # ---- Sheet 2: Action Items (all URLs) ----
        ws2 = wb.create_sheet("Action Items")
        headers = ["Priority", "Issue", "URL", "Found On / Details", "Suggestion"]
        ws2.append(headers)
        for col in range(1, 6):
            cell = ws2.cell(row=1, column=col)
            cell.fill = navy
            cell.font = white_bold
        for w in dash["works"]:
            for u in w["urls"]:
                ws2.append([w["severity"].upper(), w["label"], u["url"],
                            u.get("detail", ""), w["suggestion"]])
        for r in range(2, ws2.max_row + 1):
            for col in range(1, 6):
                ws2.cell(row=r, column=col).font = body
                ws2.cell(row=r, column=col).alignment = Alignment(vertical="top",
                                                                  wrap_text=(col in (4, 5)))
            ws2.cell(row=r, column=1).fill = sev_fill.get(
                ws2.cell(row=r, column=1).value.lower() if ws2.cell(row=r, column=1).value else "low",
                sev_fill["low"])
        widths = [10, 40, 58, 62, 62]
        for i, wd in enumerate(widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = wd
        ws2.freeze_panes = "A2"
        if ws2.max_row > 1:
            ws2.auto_filter.ref = f"A1:E{ws2.max_row}"

        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        domain = urlparse(snap["start_url"]).netloc.replace(":", "_") or "site"
        fname = f"{domain}_dashboard_{time.strftime('%Y-%m-%d')}.xlsx"
        self.send_response(200)
        self.send_header("Content-Type",
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        if self.path == "/stop":
            with STATE.lock:
                STATE.stop_flag = True
            self._json({"ok": True})
            return
        if self.path == "/search":
            try:
                payload = json.loads(self.rfile.read(length))
                q = str(payload.get("q", ""))[:200]
                mode = payload.get("mode", "contains")
                if mode not in ("contains", "not"):
                    mode = "contains"
            except Exception:
                self._json({"error": "bad request"}, 400)
                return
            if len(q) < 2:
                self._json({"results": []})
                return
            self._json({"results": custom_search(q, mode)})
            return
        if self.path != "/crawl":
            self._json({"error": "not found"}, 404)
            return
        try:
            payload = json.loads(self.rfile.read(length))
            url = payload.get("url", "").strip()
            max_pages = min(int(payload.get("max_pages", 500)), 10000)
            mode = payload.get("mode", "spider")
            if mode not in ("spider", "sitemap"):
                mode = "spider"
        except Exception:
            self._json({"error": "bad request"}, 400)
            return
        if not url:
            self._json({"error": "no url"}, 400)
            return
        if not url.startswith("http"):
            url = "https://" + url
        url = normalize_url(url)

        with STATE.lock:
            if STATE.running:
                self._json({"error": "crawl already running"}, 409)
                return
            STATE.reset()
            STATE.running = True
            STATE.start_url = url

        threading.Thread(target=crawl_worker, args=(url, max_pages, WORKERS, mode),
                         daemon=True).start()
        self._json({"ok": True})


def main():
    server = ThreadingHTTPServer(("127.0.0.1", PORT), Handler)
    url = f"http://localhost:{PORT}"
    print(f"\n  {TOOL_NAME} v{VERSION} — the free unlimited SEO crawler")
    print(f"  {TOOL_URL}")
    print(f"\n  Running at {url}")
    print("  Press Ctrl+C to stop.\n")
    threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")


if __name__ == "__main__":
    main()
